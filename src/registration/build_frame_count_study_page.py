"""Build site/frame_count_study.html - how many photos does the reconstruction actually
need? Two objects, each with its own nested, HAND-PICKED coverage-diversity subsets
(manually ranked by the user - see each object's pool_*/frame_list.txt):

  - information_sign_002_test_1: N = 25/50/75/100 x 3 methods -
        COLMAP (exp_109-111, exp_123), MASt3R-GA/swin-8 (exp_112-114, exp_124),
        MASt3R-GA/logwin-7 (exp_125-128, config/mast3r_ga_logwin.yaml - same model,
        different scenegraph construction for the global aligner)
  - bollard_003_test_1:          N = 15/30/45/60  x 2 methods (COLMAP exp_115-118,
        MASt3R-GA/swin-8 exp_119-122)

= 20 reconstructions total, each scored against its object's own LiDAR reference. Methods
don't have to match across objects either - the frontend derives each object's own method
list from its panels, same as it does for `sizes`. Note the two objects use DIFFERENT
absolute N (and pool sizes) - they're each internally nested/comparable, but N=50 on the
sign is not the same "amount of coverage" as N=50 would be on the bollard, so each object
is plotted on its own N axis rather than on a shared one.

(An earlier version of this page used exp_093-100 for the sign, a SIFT/RANSAC-overlap
greedy farthest-point selection from select_nested_diverse_frames.py instead of a human
ranking - superseded here in favor of the manual selection, which is what's actually used
for the rest of the study going forward.)

The question this page answers (that the per-object pages and capture_comparison don't):
holding object/method/capture-approach fixed, how does reconstruction quality scale with
the NUMBER of input images - does Accuracy/Completeness/F1@3cm keep improving, plateau, or
even degrade past some point? Accuracy and completeness are expected to saturate
differently (completeness keeps climbing as new viewpoints close coverage gaps; accuracy
can plateau early or even worsen as MVS/GA has more noisy, harder-to-triangulate points to
fuse) - showing that divergence is the point of the "basic curves" section.

X-axis: raw N, one axis per object. The two objects' N are not comparable to each other (a
small bollard and a ~2.5m sign need very different N to be "fully covered"), so they are
never drawn on a shared axis.

Pipeline per reconstruction (mirrors capture_comparison.html / the per-object pages
exactly, so numbers are comparable): aligned cloud -> density-match onto a fixed
1 cm voxel grid -> source<->target Chamfer distances -> gap-aware (DBSCAN)
Accuracy/Completeness/F1. The DBSCAN gap-exclusion is re-run live in the browser by a
tuner, so only the raw distance pools + positions are embedded here.

The page is structured as a loop over `OBJECTS`, each with its own `sizes` list (they
don't have to match across objects - the frontend indexes by each object's own sizes/
size_index, not a shared global), so adding a third object later is a data problem, not a
rewrite.

IMPORTANT: `import open3d` takes ~2 min in this venv, so this is deliberately ONE
monolithic process that imports o3d once and loops over all clouds in-process - never
shell out to per-cloud CLIs. Run it with `-u` to see progress.

Usage:
    python -u src/registration/build_frame_count_study_page.py
    python src/registration/build_frame_count_study_page.py --relayout   # HTML/JS only, seconds
"""

from __future__ import annotations

import base64
import json
import sys
from pathlib import Path

import numpy as np
from scipy.spatial import cKDTree

# open3d is imported lazily by `main()` - see --relayout below. The import alone costs
# ~2 min in this venv, and re-rendering the page from the payload already in the HTML
# (a template/JS change, no new numbers) doesn't need it at all.
o3d = None


def _load_open3d() -> None:
    global o3d
    if o3d is None:
        import open3d as _o3d
        o3d = _o3d


PROJECT_ROOT = Path(__file__).resolve().parents[2]
OUT_HTML = PROJECT_ROOT / "site" / "frame_count_study.html"
METRICS_JSONL = PROJECT_ROOT / "docs" / "tables" / "experiment_metrics.jsonl"

# --- configuration --------------------------------------------------------------

VOXEL_M = 0.01
# Voxel size used to density-match every reconstruction to the reference, in metres.
# Fixed at the 1 cm grid the LiDAR references are actually delivered on, rather than
# derived per-object from the reference's median NN-spacing. Two reasons: the delivered
# clouds are already thinned onto that grid, so the grid pitch is the real sampling
# limit; and the measured median NN-spacing sits *below* the pitch because overlapping
# scan passes leave exact-duplicate points at zero distance (bus_stop_001 reads 1.00 cm
# as delivered, 1.41 cm once duplicates are dropped). A fixed voxel also keeps the
# downsampling identical across objects and methods.

FLOOR_CM = 3.0          # split below/candidate here; also the min far_threshold slider value
EMBED_CAP = 6000        # max points embedded per below/candidate pool per panel
TARGET_CAP = 12000      # max reference points embedded per object (refs are small anyway)
HIST_MAX_CM = 8.0       # reference NN-spacing histogram x-range
HIST_BINS = 40

# Spatial block bootstrap for 95% CIs on Accuracy/Completeness/F1@3cm - see
# build_capture_comparison_page.py's identical note: neighbouring points of a
# density-matched cloud are not independent, so blocks (not points) are the resampling unit.
B_BOOT = 2000
THRESHOLDS_CM = [3.0, 5.0, 10.0]
# Acc/Comp/F1 are exported at all three. The DBSCAN gap mask keys off far_threshold, not the
# metric threshold, so all three read off the same kept/excluded split. 3cm stays primary: the
# bootstrap CIs and the pairwise test below are computed there.
BLOCK_CM = 5.0
BOOT_SEED = 123

METHOD_LABEL = {"colmap": "COLMAP", "mast3r_ga": "MASt3R-GA", "mast3r_ga_logwin7": "MASt3R-GA (logwin-7)"}

OBJECTS = {
    "information_sign_002": {
        "title": "information_sign_002 (manual frame selection)",
        "shape": "information sign (~2.5 m)",
        "ref_path": PROJECT_ROOT / "data/lidar/information_sign_002/information_sign_002_no_floor_centered.ply",
        # Retuned 2026-09-04 (was ft7/eps3/mp5) after the downsample fix moved this object
        # onto the fixed 1cm grid - the user rechecked it live in tuner.html. Keep in sync
        # with build_object_page.py's "information_sign" entry and build_tuner_page.py.
        "dbscan": {"ft": 5.0, "eps": 2.0, "mp": 3},
        "ref_note": "Reference partly scanned (back/edges missing); gaps DBSCAN-excluded. Same LiDAR crop used for every N and every method, so it's not a confound.",
        "nested_dir": "information_sign_002_test_1/nested_manual",
        "sizes": [25, 50, 75, 100],
        # (method, size, exp_id, aligned .ply path relative to project root)
        "entries": [
            ("colmap", 25, "exp_109", "outputs/registrations/exp_109_to_lidar_is_n25/exp_109_colmap_is_002_test_1_manual_n25_scaled_aligned_to_information_sign_002_no_floor_centered.ply"),
            ("colmap", 50, "exp_110", "outputs/registrations/exp_110_to_lidar_is_n50/exp_110_colmap_is_002_test_1_manual_n50_scaled_aligned_to_information_sign_002_no_floor_centered.ply"),
            ("colmap", 75, "exp_111", "outputs/registrations/exp_111_to_lidar_is_n75/exp_111_colmap_is_002_test_1_manual_n75_scaled_aligned_to_information_sign_002_no_floor_centered.ply"),
            ("colmap", 100, "exp_123", "outputs/registrations/exp_123_to_lidar_is_n_100/exp_123_colmap_is_002_test_1_manual_n100_scaled_aligned_to_information_sign_002_no_floor_centered.ply"),
            ("mast3r_ga", 25, "exp_112", "outputs/registrations/exp_112_to_lidar_is_n25/exp_112_mast3r_ga_is_002_test_1_manual_n25_scaled_aligned_to_information_sign_002_no_floor_centered.ply"),
            ("mast3r_ga", 50, "exp_113", "outputs/registrations/exp_113_to_lidar_is_n_50/exp_113_mast3r_ga_is_002_test_1_manual_n50_scaled_aligned_to_information_sign_002_no_floor_centered.ply"),
            ("mast3r_ga", 75, "exp_114", "outputs/registrations/exp_114_to_lidar_is_n75/exp_114_mast3r_ga_is_002_test_1_manual_n75_scaled_aligned_to_information_sign_002_no_floor_centered.ply"),
            ("mast3r_ga", 100, "exp_124", "outputs/registrations/exp_124_to_lidar_is_n100/exp_124_mast3r_ga_is_002_test_1_manual_n100_scaled_aligned_to_information_sign_002_no_floor_centered.ply"),
            # MASt3R-GA with scenegraph=logwin-7 instead of the default swin-8 (config/mast3r_ga_logwin.yaml)
            ("mast3r_ga_logwin7", 25, "exp_128", "outputs/registrations/exp_128_to_lidar_information_sign_002/exp_128_mast3r_ga_is_002_test_1_manual_n25_scaled_aligned_to_information_sign_002_no_floor_centered.ply"),
            ("mast3r_ga_logwin7", 50, "exp_127", "outputs/registrations/exp_127_to_lidar_information_sign_002/exp_127_mast3r_ga_is_002_test_1_manual_n50_scaled_aligned_to_information_sign_002_no_floor_centered.ply"),
            ("mast3r_ga_logwin7", 75, "exp_126", "outputs/registrations/exp_126_to_lidar_information_sign_002/exp_126_mast3r_ga_is_002_test_1_manual_n75_scaled_aligned_to_information_sign_002_no_floor_centered.ply"),
            ("mast3r_ga_logwin7", 100, "exp_125", "outputs/registrations/exp_125_to_lidar_information_sign_002/exp_125_mast3r_ga_is_002_test_1_manual_n100_scaled_aligned_to_information_sign_002_no_floor_centered.ply"),
        ],
    },
    "bollard_003": {
        "title": "bollard_003 (manual frame selection)",
        "shape": "bollard (~1 m post)",
        "ref_path": PROJECT_ROOT / "data/lidar/bollard_003/bollard_003_no_floor_centered.ply",
        # same tuned DBSCAN defaults as bollard_003 elsewhere in the project (capture_comparison.html)
        "dbscan": {"ft": 5.0, "eps": 2.0, "mp": 3},
        "ref_note": "Reference missing a ~90° sector (DBSCAN-excluded, not counted as error). Same LiDAR crop used for every N and every method, so it's not a confound.",
        "nested_dir": "bollard_003_test_1/nested_manual",
        "sizes": [15, 30, 45, 60],
        # (method, size, exp_id, aligned .ply path relative to project root)
        "entries": [
            ("colmap", 15, "exp_115", "outputs/registrations/exp_115_to_lidar_bollard_003/exp_115_colmap_bollard_003_test_1_manual_n15_scaled_aligned_to_bollard_003_no_floor_centered.ply"),
            ("colmap", 30, "exp_116", "outputs/registrations/exp_116_to_lidar_bollard_003/exp_116_colmap_bollard_003_test_1_manual_n30_scaled_aligned_to_bollard_003_no_floor_centered.ply"),
            ("colmap", 45, "exp_117", "outputs/registrations/exp_117_to_lidar_bollard_003/exp_117_colmap_bollard_003_test_1_manual_n45_scaled_aligned_to_bollard_003_no_floor_centered.ply"),
            ("colmap", 60, "exp_118", "outputs/registrations/exp_118_to_lidar_bollard_003/exp_118_colmap_bollard_003_test_1_manual_n60_scaled_aligned_to_bollard_003_no_floor_centered.ply"),
            ("mast3r_ga", 15, "exp_119", "outputs/registrations/exp_119_to_lidar_bollard_003/exp_119_mast3r_ga_bollard_003_test_1_manual_n15_scaled_aligned_to_bollard_003_no_floor_centered.ply"),
            ("mast3r_ga", 30, "exp_120", "outputs/registrations/exp_120_to_lidar_bollard_003/exp_120_mast3r_ga_bollard_003_test_1_manual_n30_scaled_aligned_to_bollard_003_no_floor_centered.ply"),
            ("mast3r_ga", 45, "exp_121", "outputs/registrations/exp_121_to_lidar_bollard_003/exp_121_mast3r_ga_bollard_003_test_1_manual_n45_scaled_aligned_to_bollard_003_no_floor_centered.ply"),
            ("mast3r_ga", 60, "exp_122", "outputs/registrations/exp_122_to_lidar_bollard_003/exp_122_mast3r_ga_bollard_003_test_1_manual_n60_scaled_aligned_to_bollard_003_no_floor_centered.ply"),
        ],
    },
}

rng = np.random.default_rng(42)


# --- helpers --------------------------------------------------------------------

def b64f(arr) -> str:
    """base64 of a little-endian float32 array (flattened C-order)."""
    return base64.b64encode(np.ascontiguousarray(arr, dtype="<f4").ravel().tobytes()).decode("ascii")


def subsample(idx_count: int, cap: int) -> np.ndarray:
    if idx_count <= cap:
        return np.arange(idx_count)
    return np.sort(rng.choice(idx_count, cap, replace=False))


def nn_spacing_all(points: np.ndarray) -> np.ndarray:
    """Distance from each point to its nearest OTHER point in the same cloud."""
    tree = cKDTree(points)
    d, _ = tree.query(points, k=2)
    return d[:, 1]


def reg_rates_from_experiments_yaml() -> dict[str, float]:
    """Registration rate per exp_id, parsed from config/experiments.yaml's "Registered
    images: N/M" log line, as a fraction.

    Fallback for reg_rates_from_metrics(): every experiment in this ablation (exp_109-128)
    predates metrics.py, so docs/tables/experiment_metrics.jsonl has no row for any of them
    and the column came out empty for all 20. experiments.yaml carries the same figure.
    """
    import re as _re
    path = PROJECT_ROOT / "config" / "experiments.yaml"
    if not path.exists():
        return {}
    out: dict[str, float] = {}
    for m in _re.finditer(r"^  (exp_\d+):\n(.*?)(?=^  exp_|\Z)", path.read_text(), _re.S | _re.M):
        got = _re.search(r"Registered images: (\d+)/(\d+)", m.group(2))
        if got and int(got.group(2)):
            out[m.group(1)] = int(got.group(1)) / int(got.group(2))
    return out


def reg_rates_from_metrics() -> dict[str, float]:
    out: dict[str, float] = {}
    if not METRICS_JSONL.exists():
        return out
    for line in METRICS_JSONL.read_text().splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            r = json.loads(line)
        except json.JSONDecodeError:
            continue
        eid = r.get("exp_id")
        rate = r.get("registration_rate")
        if eid and rate is not None:
            out[eid] = rate
    return out


def f_score(p: float, r: float) -> float:
    return 0.0 if (p + r) == 0 else 2 * p * r / (p + r)


# --- spatial block bootstrap ----------------------------------------------------

def _block_parts(points_m: np.ndarray, indicator: np.ndarray, block_m: float):
    """Per spatial block (a `block_m`-sized voxel): how many points fall in it and how
    many of those satisfy `indicator` (e.g. distance <= 3cm). These per-block sums are
    what the bootstrap resamples, so within-block correlation is preserved."""
    if len(points_m) == 0:
        return np.array([]), np.array([]), 0
    q = np.floor(points_m / block_m).astype(np.int64)
    _, inv = np.unique(q, axis=0, return_inverse=True)
    nb = int(inv.max()) + 1
    within = np.bincount(inv, weights=indicator.astype(float), minlength=nb)
    total = np.bincount(inv, minlength=nb).astype(float)
    return within, total, nb


def bootstrap_draws(acc_pts, acc_ind, comp_pts, comp_ind, block_m, B, rng):
    """B block-bootstrap draws each of Accuracy@3cm, Completeness@3cm and F1@3cm (%).
    Accuracy blocks (kept source points) and completeness blocks (target points) are
    resampled independently each iteration."""
    aw, at, anb = _block_parts(acc_pts, acc_ind, block_m)
    cw, ct, cnb = _block_parts(comp_pts, comp_ind, block_m)
    if anb == 0 or cnb == 0:
        return None, None, None, anb, cnb
    ai = rng.integers(0, anb, size=(B, anb))
    acc_b = aw[ai].sum(1) / at[ai].sum(1)
    ci = rng.integers(0, cnb, size=(B, cnb))
    comp_b = cw[ci].sum(1) / ct[ci].sum(1)
    denom = acc_b + comp_b
    f1_b = np.where(denom > 0, 2 * acc_b * comp_b / denom, 0.0) * 100.0
    return acc_b * 100.0, comp_b * 100.0, f1_b, anb, cnb


def ci95(draws):
    if draws is None:
        return float("nan"), float("nan")
    lo, hi = np.percentile(draws, [2.5, 97.5])
    return float(lo), float(hi)


# --- main ------------------------------------------------------------------------

def load_reference(path) -> o3d.geometry.PointCloud:
    """Read a reference (LiDAR) cloud and drop exact-duplicate points before any
    metric touches it.

    The delivered clouds carry a lot of them - 10-26% of the points on these objects
    (bus_stop_001 41337 -> 31692, flashlight_003 22637 -> 16690, bollard_003 2818 ->
    2526) - left over where overlapping scan passes cover the same surface twice, at
    byte-identical coordinates. They are not extra information, but they do skew the
    metrics: completeness is averaged over target points, so a duplicated point is
    counted as many times as it appears, quietly weighting the score toward whatever
    the scanner happened to pass twice. Distances themselves are unaffected (a
    duplicate is its own nearest neighbour), so this only removes the double-counting.
    """
    pcd = o3d.io.read_point_cloud(str(path))
    pts = np.asarray(pcd.points)
    _, first_idx = np.unique(pts, axis=0, return_index=True)
    if len(first_idx) < len(pts):
        # select_by_index (rather than rebuilding from the array) so colors/normals survive
        pcd = pcd.select_by_index(np.sort(first_idx).tolist())
        print(f"       deduplicated reference: {len(pts)} -> {len(first_idx)} points "
              f"({100 * (1 - len(first_idx) / len(pts)):.1f}% exact duplicates removed)", flush=True)
    return pcd


def main() -> None:
    _load_open3d()
    reg_rates = reg_rates_from_metrics()
    for _e, _v in reg_rates_from_experiments_yaml().items():
        reg_rates.setdefault(_e, _v)   # jsonl wins where present; yaml fills the gaps

    objects_data: list[dict] = []
    panels: dict[str, dict] = {}
    boot_rng = np.random.default_rng(BOOT_SEED)
    f1_draws: dict[str, np.ndarray] = {}    # panel key -> B F1@3cm draws, for the pairwise-N test
    acc_draws: dict[str, np.ndarray] = {}   # same for accuracy: the thesis claim is about accuracy, not F1

    for obj_id, cfg in OBJECTS.items():
        print(f"[ref] {obj_id}: loading {cfg['ref_path'].name}", flush=True)
        ref = load_reference(cfg["ref_path"])
        rpts = np.asarray(ref.points)
        spacing = nn_spacing_all(rpts)
        median_spacing = float(np.median(spacing))
        height_m = float(rpts[:, 2].max() - rpts[:, 2].min())
        print(f"       {len(rpts)} pts, median NN spacing = {median_spacing*100:.3f} cm, "
              f"height = {height_m:.2f} m", flush=True)

        spacing_cm = spacing * 100.0
        edges = np.linspace(0.0, HIST_MAX_CM, HIST_BINS + 1)
        counts, _ = np.histogram(np.clip(spacing_cm, 0, HIST_MAX_CM), bins=edges)
        overflow_pct = float(np.mean(spacing_cm > HIST_MAX_CM) * 100.0)

        tsub = subsample(len(rpts), TARGET_CAP)
        target_pos_embed = rpts[tsub]

        obj_entry = {
            "id": obj_id,
            "title": cfg["title"],
            "shape": cfg["shape"],
            "ref_note": cfg["ref_note"],
            "ref_spacing_cm": round(median_spacing * 100, 4),
            "height_m": round(height_m, 2),
            "dbscan": cfg["dbscan"],
            "hist": {
                "n": int(len(rpts)),
                "mean_cm": float(spacing_cm.mean()),
                "median_cm": float(np.median(spacing_cm)),
                "p95_cm": float(np.percentile(spacing_cm, 95)),
                "max_cm": float(spacing_cm.max()),
                "overflow_pct": overflow_pct,
                "counts": counts.astype(int).tolist(),
                "bin_edges_cm": [round(e, 4) for e in edges.tolist()],
            },
            "target_pos": b64f(target_pos_embed),
            "n_target_total": int(len(rpts)),
            "sizes": cfg["sizes"],
            "n_max": max(cfg["sizes"]),
            "panels": [],
        }

        for method, size, exp_id, rel in cfg["entries"]:
            key = f"{obj_id}__{method}__{size}"
            src_path = PROJECT_ROOT / rel
            print(f"[panel] {key}  ({exp_id})  <- {src_path.name}", flush=True)
            src = o3d.io.read_point_cloud(str(src_path))
            raw_points = len(src.points)

            matched = src.voxel_down_sample(VOXEL_M)
            mpts = np.asarray(matched.points)
            n_matched = len(mpts)

            d_s2t = np.asarray(matched.compute_point_cloud_distance(ref))  # meters
            d_s2t_cm = d_s2t * 100.0
            d_t2s_cm = np.asarray(ref.compute_point_cloud_distance(matched)) * 100.0

            below_mask = d_s2t_cm <= FLOOR_CM
            cand_mask = ~below_mask
            n_below_true = int(below_mask.sum())
            n_cand_true = int(cand_mask.sum())

            below_idx = np.where(below_mask)[0]
            cand_idx = np.where(cand_mask)[0]
            below_sel = below_idx[subsample(len(below_idx), EMBED_CAP)]
            cand_sel = cand_idx[subsample(len(cand_idx), EMBED_CAP)]

            dcfg = cfg["dbscan"]
            far_m = dcfg["ft"] / 100.0
            cand_far_idx = np.where(d_s2t > far_m)[0]
            gap_mask = np.zeros(n_matched, dtype=bool)
            if len(cand_far_idx) > 0:
                far_pcd = o3d.geometry.PointCloud()
                far_pcd.points = o3d.utility.Vector3dVector(mpts[cand_far_idx])
                labels = np.array(far_pcd.cluster_dbscan(eps=dcfg["eps"] / 100.0, min_points=dcfg["mp"]))
                gap_mask[cand_far_idx[labels >= 0]] = True
            kept = ~gap_mask
            n_excluded = int(gap_mask.sum())
            d_kept_cm = d_s2t_cm[kept]
            by_threshold = {}
            for _t in THRESHOLDS_CM:
                _a = float(np.mean(d_kept_cm <= _t)) if d_kept_cm.size else 0.0
                _c = float(np.mean(d_t2s_cm <= _t)) if d_t2s_cm.size else 0.0
                by_threshold[f"{_t:g}cm"] = (_a, _c, f_score(_a, _c))
            acc_3, comp_3 = by_threshold["3cm"][0], by_threshold["3cm"][1]
            inliers_3 = d_kept_cm[d_kept_cm <= 3.0]
            inlier_rmse_3 = float(np.sqrt(np.mean(inliers_3 ** 2))) if inliers_3.size else float("nan")
            f1_3 = f_score(acc_3, comp_3)
            acc_median = float(np.median(d_kept_cm)) if d_kept_cm.size else float("nan")
            comp_median = float(np.median(d_t2s_cm)) if d_t2s_cm.size else float("nan")

            acc_b, comp_b, f1_b, n_blocks_acc, n_blocks_comp = bootstrap_draws(
                mpts[kept], d_kept_cm <= 3.0, rpts, d_t2s_cm <= 3.0, BLOCK_CM / 100.0, B_BOOT, boot_rng,
            )
            f1_draws[key] = f1_b
            acc_draws[key] = acc_b
            acc_ci_lo, acc_ci_hi = ci95(acc_b)
            comp_ci_lo, comp_ci_hi = ci95(comp_b)
            f1_ci_lo, f1_ci_hi = ci95(f1_b)

            print(f"          N={size:4d} {METHOD_LABEL[method]:10s} matched={n_matched:8d} excluded={n_excluded:6d}  "
                  f"F1@3={f1_3*100:5.1f}% [{f1_ci_lo:.1f},{f1_ci_hi:.1f}]  acc@3={acc_3*100:5.1f}%  "
                  f"comp@3={comp_3*100:5.1f}%  acc_med={acc_median:.2f}cm  comp_med={comp_median:.2f}cm", flush=True)

            panels[key] = {
                "object": obj_id,
                "method": method,
                "size": size,
                "size_index": cfg["sizes"].index(size),  # rank (0=smallest..3=largest) within THIS object's own sizes
                "exp_id": exp_id,
                "label": f"{METHOD_LABEL[method]} · N={size}",
                "raw_points": raw_points,
                "matched_points": n_matched,
                "reg_rate": reg_rates.get(exp_id),
                "accuracy_median_cm": round(acc_median, 3),
                "completeness_median_cm": round(comp_median, 3),
                "below_pos": b64f(mpts[below_sel]),
                "below_dist_cm": b64f(d_s2t_cm[below_sel]),
                "candidate_pos": b64f(mpts[cand_sel]),
                "candidate_dist_cm": b64f(d_s2t_cm[cand_sel]),
                "target_dist_cm": b64f(d_t2s_cm),
                "n_source_total": n_matched,
                "n_below_true": n_below_true,
                "n_candidates_true": n_cand_true,
                "n_target_total": int(len(rpts)),
                "n_over_nmax": round(size / max(cfg["sizes"]), 4),
                "default": {
                    "f1_3cm": round(f1_3 * 100, 2),
                    "acc_3cm": round(acc_3 * 100, 2),
                    "comp_3cm": round(comp_3 * 100, 2),
                    **{f"{m}_{k}": round(v * 100, 2)
                       for k, (a, c, f) in by_threshold.items()
                       for m, v in (("acc", a), ("comp", c), ("f1", f))
                       if k != "3cm"},
                    "f1_delta_10_3": round((by_threshold["10cm"][2] - by_threshold["3cm"][2]) * 100, 2),
                    "inlier_rmse_3cm": round(inlier_rmse_3, 3),
                    "n_excluded": n_excluded,
                    "f1_ci_lo": round(f1_ci_lo, 2), "f1_ci_hi": round(f1_ci_hi, 2),
                    "acc_ci_lo": round(acc_ci_lo, 2), "acc_ci_hi": round(acc_ci_hi, 2),
                    "comp_ci_lo": round(comp_ci_lo, 2), "comp_ci_hi": round(comp_ci_hi, 2),
                    "n_blocks_acc": int(n_blocks_acc), "n_blocks_comp": int(n_blocks_comp),
                },
            }
            obj_entry["panels"].append(key)

        objects_data.append(obj_entry)

    # ----- pairwise test between N levels -------------------------------------------
    # The per-point CIs already on the page say how precise each N is on its own; they do
    # not say whether N=100 actually beats N=25. Differencing the paired bootstrap draws
    # does: the same resampled blocks feed both sides, so the block-to-block variation
    # cancels and what is left is the difference attributable to frame count. A CI that
    # spans 0 means the two frame counts are indistinguishable at this sample size.
    n_significance = []
    for obj in objects_data:
        obj_id = obj["id"]
        methods = sorted({panels[k]["method"] for k in obj["panels"]})
        sizes = OBJECTS[obj_id]["sizes"]
        for method in methods:
            avail = [n for n in sizes if f"{obj_id}__{method}__{n}" in f1_draws]
            for i, n_lo in enumerate(avail):
                for n_hi in avail[i + 1:]:
                    for metric, store in (("F1", f1_draws), ("accuracy", acc_draws)):
                        a = store.get(f"{obj_id}__{method}__{n_lo}")
                        b = store.get(f"{obj_id}__{method}__{n_hi}")
                        if a is None or b is None:
                            continue
                        d = np.asarray(b) - np.asarray(a)      # larger N minus smaller N
                        lo, hi = (float(x) for x in np.percentile(d, [2.5, 97.5]))
                        n_significance.append({
                            "object": obj_id, "method": method, "metric": metric,
                            "pair": f"N={n_hi} - N={n_lo}",
                            "n_lo": n_lo, "n_hi": n_hi,
                            "delta": round(float(d.mean()), 2),
                            "ci_lo": round(lo, 2), "ci_hi": round(hi, 2),
                            "includes_zero": bool(lo <= 0 <= hi),
                        })

    print("\nDoes adding frames change anything? (paired bootstrap, smallest vs each larger N)", flush=True)
    for r in n_significance:
        if r["n_lo"] != min(OBJECTS[r["object"]]["sizes"]):
            continue
        verdict = "within noise" if r["includes_zero"] else "RESOLVABLE"
        print(f"  {r['object']:<22}{METHOD_LABEL[r['method']]:<22}{r['metric']:<9}{r['pair']:<18}"
              f"Δ={r['delta']:+6.2f}  95% CI=[{r['ci_lo']:+.2f},{r['ci_hi']:+.2f}]  {verdict}", flush=True)

    data = {
        "floor_cm": FLOOR_CM,
        "n_significance": n_significance,
        "objects": objects_data,
        "panels": panels,
        "method_label": METHOD_LABEL,
        "bootstrap": {"n_draws": B_BOOT, "block_cm": BLOCK_CM},
    }

    html = build_html(data)
    OUT_HTML.parent.mkdir(parents=True, exist_ok=True)
    OUT_HTML.write_text(html, encoding="utf-8")
    size_mb = OUT_HTML.stat().st_size / (1024 * 1024)
    print(f"\nWrote {OUT_HTML.relative_to(PROJECT_ROOT)}  ({size_mb:.2f} MB)", flush=True)

    summary_path = PROJECT_ROOT / "docs" / "tables" / "frame_count_study_summary.json"
    summary = [
        {
            "object": p["object"], "method": p["method"], "size": p["size"],
            "exp_id": p["exp_id"], "reg_rate": p["reg_rate"],
            "raw_points": p["raw_points"], "matched_points": p["matched_points"],
            "n_over_nmax": p["n_over_nmax"],
            "accuracy_median_cm": p["accuracy_median_cm"],
            "completeness_median_cm": p["completeness_median_cm"],
            **p["default"],
        }
        for p in panels.values()
    ]
    summary_path.write_text(json.dumps(summary, indent=2))
    print(f"Wrote {summary_path.relative_to(PROJECT_ROOT)}", flush=True)

    xlsx_path = PROJECT_ROOT / "docs" / "tables" / "frame_count_study_summary.xlsx"
    write_summary_xlsx(summary, n_significance, xlsx_path)
    print(f"Wrote {xlsx_path.relative_to(PROJECT_ROOT)}", flush=True)


def write_summary_xlsx(summary: list[dict], n_significance: list[dict], path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    def head(ws, headers):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F3864")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thresh = [f"{t:g}cm" for t in THRESHOLDS_CM]
    wb = Workbook()
    ws = wb.active
    ws.title = "frame_count_study"
    headers = ["object", "method", "N", "exp_id", "reg-rate (%)", "raw_points", "matched_points",
               "accuracy median (cm)", "completeness median (cm)", "inlier RMSE@3cm (mm)"]
    for k in thresh:
        headers += [f"accuracy@{k} (%)", f"completeness@{k} (%)", f"F1@{k} (%)"]
    headers += ["ΔF1@10-3cm (pp)", "F1@3cm CI lo", "F1@3cm CI hi",
                "Acc@3cm CI lo", "Acc@3cm CI hi", "Comp@3cm CI lo", "Comp@3cm CI hi", "excluded as gap"]
    head(ws, headers)

    for row in sorted(summary, key=lambda r: (r["object"], r["method"], r["size"])):
        vals = [row["object"], row["method"], row["size"], row["exp_id"],
                (round(row["reg_rate"] * 100, 1) if row.get("reg_rate") is not None else None),
                row["raw_points"], row["matched_points"],
                row["accuracy_median_cm"], row["completeness_median_cm"],
                # mm, matching summary_all_objects_accuracy_f1.xlsx; it was cm here alone
                round(row["inlier_rmse_3cm"] * 10, 2)]
        for k in thresh:
            vals += [row.get(f"acc_{k}"), row.get(f"comp_{k}"), row.get(f"f1_{k}")]
        vals += [row.get("f1_delta_10_3"), row["f1_ci_lo"], row["f1_ci_hi"],
                 row["acc_ci_lo"], row["acc_ci_hi"], row["comp_ci_lo"], row["comp_ci_hi"],
                 row["n_excluded"]]
        ws.append(vals)
    for col in ws.columns:
        width = max(len(str(c.value)) if c.value is not None else 0 for c in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(width, 20)
    ws.freeze_panes = "E2"

    # sheet 2: does adding frames actually change anything?
    ws2 = wb.create_sheet("significance")
    ws2.append([f"Paired block-bootstrap difference between frame counts, "
                f"{B_BOOT} draws, {BLOCK_CM:g} cm blocks. CI spanning 0 = not resolvable."])
    ws2["A1"].font = Font(italic=True, color="585D54")
    ws2.append([])
    ws2.append(["object", "method", "metric", "pair", "Δ@3cm (pp)", "95% CI lo", "95% CI hi", "resolvable?"])
    for c in ws2[3]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F3864")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in n_significance:
        ws2.append([r["object"], METHOD_LABEL.get(r["method"], r["method"]), r.get("metric", "F1"), r["pair"],
                    r["delta"], r["ci_lo"], r["ci_hi"],
                    "no - CI spans 0" if r["includes_zero"] else "yes"])
    for row in ws2.iter_rows(min_row=4, min_col=8, max_col=8):
        for c in row:
            if c.value == "yes":
                c.font = Font(bold=True, color="0D8054")
    for col, w in zip("ABCDEFGH", (22, 22, 10, 18, 14, 12, 12, 17)):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A4"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def build_html(data: dict) -> str:
    return wrap_html(json.dumps(data).replace("</", "<\\/"))


def wrap_html(payload: str) -> str:
    return HTML_HEAD + f'\n<script type="application/json" id="page-data">{payload}</script>\n' + MAIN_JS + HTML_TAIL


def relayout() -> None:
    """Re-render site/frame_count_study.html from the payload already embedded in it.

    The page is HTML_HEAD + <the data> + MAIN_JS + HTML_TAIL, and the data is the only
    part that costs anything to produce (20 clouds through open3d). So a template-only
    change - new chart, restyled table, reworded copy - can reuse the payload verbatim
    and skip the numeric work entirely. Use this ONLY when the payload itself is
    unchanged: any new field the template reads has to come from a full rebuild.
    """
    import re

    html = OUT_HTML.read_text(encoding="utf-8")
    m = re.search(r'<script type="application/json" id="page-data">(.*?)</script>', html, re.S)
    if not m:
        sys.exit(f"no embedded payload in {OUT_HTML} - run a full rebuild instead")
    OUT_HTML.write_text(wrap_html(m.group(1)), encoding="utf-8")
    size_mb = OUT_HTML.stat().st_size / (1024 * 1024)
    print(f"Re-rendered {OUT_HTML.relative_to(PROJECT_ROOT)} from its existing payload ({size_mb:.2f} MB)")


# HTML_HEAD / MAIN_JS / HTML_TAIL are defined in the template module below to keep this
# file readable; they are imported at module load.
from _frame_count_page_template import HTML_HEAD, MAIN_JS, HTML_TAIL  # noqa: E402


if __name__ == "__main__":
    if "--relayout" in sys.argv[1:]:
        sys.exit(relayout())
    sys.exit(main())
