"""Build site/capture_comparison.html - a self-contained, interactive comparison of
CAPTURE APPROACHES (how the object was filmed) across 2 objects x 3 approaches x 2
methods = 12 reconstructions, all scored against their LiDAR reference.

The question this page answers (that the per-object pages don't): holding object and
reconstruction method fixed, how much does the *camera capture strategy* change the
result, and is one strategy a robust winner across both objects and both methods?

Capture approaches ("tests"):
  1 (test_1): close-range views PLUS distant views of the whole object - mixed scales.
  2 (test_2): close-range only - close-range views, no distant views.
  3 (test_3): distant only - object always fully in frame, no close-range views.

Pipeline per reconstruction (mirrors the per-object pages exactly, so numbers are
comparable): aligned cloud -> density-match onto a fixed 1 cm voxel grid ->
source<->target Chamfer distances -> gap-aware (DBSCAN) Accuracy/Completeness/F1. The
DBSCAN gap-exclusion is re-run live in the browser by a per-object tuner, so only the
raw distance pools + positions are embedded here (same data layout the bus_stop page
uses).

IMPORTANT: `import open3d` takes ~2 min in this venv, so this is deliberately ONE
monolithic process that imports o3d once and loops over all 12 clouds in-process -
never shell out to the per-cloud CLIs. Run it with `-u` to see progress.

Usage:
    python -u src/registration/build_capture_comparison_page.py
"""

from __future__ import annotations

import base64
import json
import sys
from pathlib import Path

import numpy as np
import open3d as o3d
from scipy.spatial import cKDTree

PROJECT_ROOT = Path(__file__).resolve().parents[2]
OUT_HTML = PROJECT_ROOT / "site" / "capture_comparison.html"
METRICS_JSONL = PROJECT_ROOT / "docs" / "tables" / "experiment_metrics.jsonl"

# --- configuration --------------------------------------------------------------

VOXEL_M = 0.01
# Voxel size used to density-match every reconstruction to the reference, in metres.
# Fixed at the 1 cm grid the LiDAR references are actually delivered on, rather than
# derived per-object from the reference's median NN-spacing. Two reasons: the delivered
# clouds are already thinned onto that grid, so the grid pitch is the real sampling
# limit; and the measured median NN-spacing sits *below* the pitch because overlapping
# scan passes leave exact-duplicate points at zero distance (bus_stop_001 reads 1.00 cm
# as delivered, 1.41 cm once duplicates are dropped). A fixed voxel also keeps the
# downsampling identical across objects and methods.

FLOOR_CM = 3.0          # split below/candidate here; also the min far_threshold slider value
THRESHOLDS_CM = [3.0, 5.0, 10.0]
# Acc/Comp/F1 are reported at all three. The DBSCAN gap mask depends on far_threshold, not on
# the metric threshold, so all three are read off the same kept/excluded split - only the <=
# comparison changes. 3cm stays primary (it is what the bootstrap CI and the charts use);
# 5/10cm exist so a weak result can be read as 'offset' (closes as the threshold widens) vs
# 'absent' (stays low even at 10cm), which is what the delta column makes explicit.
EMBED_CAP = 6000        # max points embedded per below/candidate pool per panel
TARGET_CAP = 12000      # max reference points embedded per object (refs are small anyway)
HIST_MAX_CM = 8.0       # reference NN-spacing histogram x-range
HIST_BINS = 40

# Spatial block bootstrap for a 95% CI on F1@3cm. A plain per-point (binomial) CI
# would be dishonestly narrow: neighbouring points of a density-matched cloud are not
# independent (the source->target distance field is spatially smooth), so the effective
# sample size is the number of independent surface patches, not the raw point count.
# Resampling ~BLOCK_CM-sized spatial blocks (with replacement) respects that correlation
# and yields an honest, wider interval. BLOCK_CM is a few x the point spacing and the 3cm
# threshold, so a block is one roughly-independent patch of surface.
B_BOOT = 2000
BLOCK_CM = 5.0
BOOT_SEED = 123

REFS = {
    "bollard_003": PROJECT_ROOT / "data/lidar/bollard_003/bollard_003_no_floor_centered.ply",
    "information_sign_002": PROJECT_ROOT / "data/lidar/information_sign_002/information_sign_002_no_floor_centered.ply",
}

OBJECT_META = {
    "bollard_003": {
        "title": "bollard_003",
        "shape": "bollard (~1 m post)",
        "dbscan": {"ft": 5.0, "eps": 2.0, "mp": 3},
        "ref_note": "Reference missing a ~90° sector (DBSCAN-excluded, not counted as error).",
    },
    "information_sign_002": {
        "title": "information_sign_002",
        "shape": "information sign (~2.5 m)",
        # Retuned 2026-09-04 (was ft7/eps3/mp5) after the downsample fix moved this object
        # onto the fixed 1cm grid - the user rechecked it live in tuner.html. Keep in sync
        # with build_object_page.py's "information_sign" entry and build_tuner_page.py.
        "dbscan": {"ft": 5.0, "eps": 2.0, "mp": 3},
        "ref_note": "Reference partly scanned (back/edges missing); gaps DBSCAN-excluded.",
    },
}

METHOD_LABEL = {"colmap": "COLMAP", "mast3r_ga": "MASt3R-GA"}

# (object, method, approach, exp_id, aligned .ply path relative to project root)
ENTRIES = [
    ("bollard_003", "colmap", 1, "exp_087", "outputs/registrations/exp_087_to_lidar_bollard_003_test_1/exp_087_colmap_bollard_003_test_1_scaled.ply"),
    ("bollard_003", "colmap", 2, "exp_088", "outputs/registrations/exp_088_to_lidar_bollard_003_test_2/exp_088_colmap_bollard_003_test_2.ply"),
    ("bollard_003", "colmap", 3, "exp_089", "outputs/registrations/exp_089_to_lidar_bollard_003_test_3/exp_089_colmap_bollard_003_test_3.ply"),
    ("bollard_003", "mast3r_ga", 1, "exp_090", "outputs/registrations/exp_090_to_lidar_bollard_003_test_1/exp_090_mast3r_ga_bollard_003_test_1.ply"),
    ("bollard_003", "mast3r_ga", 2, "exp_091", "outputs/registrations/exp_091_to_lidar_bollard_003_test_2/exp_091_mast3r_ga_bollard_003_test_2.ply"),
    ("bollard_003", "mast3r_ga", 3, "exp_092", "outputs/registrations/exp_092_to_lidar_bollard_003_test_3/exp_092_mast3r_ga_bollard_003_test_3.ply"),
    ("information_sign_002", "colmap", 1, "exp_081", "outputs/registrations/exp_081_to_lidar_is_002_test_1/exp_081_colmap_is_002_test_1.ply"),
    ("information_sign_002", "colmap", 2, "exp_082", "outputs/registrations/exp_082_to_lidar_is_002_test_2/exp_082_colmap_is_002_test_2.ply"),
    # exp_083 is COLMAP test_3 (verified: 716317 verts == exp_083 colmap source); the on-disk
    # filename says "mast3r" but that is a misnomer, confirmed by the user.
    ("information_sign_002", "colmap", 3, "exp_083", "outputs/registrations/exp_083_to_lidar_is_test_3/exp_083_mast3r_is_002_test_3.ply"),
    ("information_sign_002", "mast3r_ga", 1, "exp_084", "outputs/registrations/exp_084_to_lidar_is_002_test_1/exp_084_mast3r_is_002_test_1.ply"),
    ("information_sign_002", "mast3r_ga", 2, "exp_085", "outputs/registrations/exp_085_to_lidar_is_002_test_2/exp_085_mast3r_is_002_test_2.ply"),
    ("information_sign_002", "mast3r_ga", 3, "exp_086", "outputs/registrations/exp_086_to_lidar_is_002_test_3/exp_086_mast3r_is_002_test_3.ply"),
]

APPROACH_LABEL = {
    1: "T1 · close-range + distant",
    2: "T2 · close-range only",
    3: "T3 · distant only",
}

rng = np.random.default_rng(42)


# --- helpers --------------------------------------------------------------------

def b64f(arr) -> str:
    """base64 of a little-endian float32 array (flattened C-order)."""
    return base64.b64encode(np.ascontiguousarray(arr, dtype="<f4").ravel().tobytes()).decode("ascii")


def subsample(idx_count: int, cap: int) -> np.ndarray:
    if idx_count <= cap:
        return np.arange(idx_count)
    return np.sort(rng.choice(idx_count, cap, replace=False))


def nn_spacing_all(points: np.ndarray) -> np.ndarray:
    """Distance from each point to its nearest OTHER point in the same cloud."""
    tree = cKDTree(points)
    d, _ = tree.query(points, k=2)
    return d[:, 1]


def reg_rates_from_experiments_yaml() -> dict[str, float]:
    """Registration rate per exp_id, parsed from config/experiments.yaml's "Registered
    images: N/M" log line.

    Fallback for reg_rates_from_metrics(): docs/tables/experiment_metrics.jsonl lost its
    rows for exp_081-092 (overwritten 2026-08-31; only the performance-study experiments
    were recoverable), so for these twelve the jsonl reports nothing. experiments.yaml is
    intact and carries the same figure, so the column need not stay empty.
    """
    import re as _re
    path = PROJECT_ROOT / "config" / "experiments.yaml"
    if not path.exists():
        return {}
    out: dict[str, float] = {}
    for m in _re.finditer(r"^  (exp_\d+):\n(.*?)(?=^  exp_|\Z)", path.read_text(), _re.S | _re.M):
        got = _re.search(r"Registered images: (\d+)/(\d+)", m.group(2))
        if got and int(got.group(2)):
            out[m.group(1)] = round(int(got.group(1)) / int(got.group(2)) * 100, 1)
    return out


def reg_rates_from_metrics() -> dict[str, float]:
    out: dict[str, float] = {}
    if not METRICS_JSONL.exists():
        return out
    for line in METRICS_JSONL.read_text().splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            r = json.loads(line)
        except json.JSONDecodeError:
            continue
        eid = r.get("exp_id")
        rate = r.get("registration_rate")
        if eid and rate is not None:
            out[eid] = rate
    return out


def f_score(p: float, r: float) -> float:
    return 0.0 if (p + r) == 0 else 2 * p * r / (p + r)


# --- spatial block bootstrap ----------------------------------------------------

def _block_parts(points_m: np.ndarray, indicator: np.ndarray, block_m: float):
    """Per spatial block (a `block_m`-sized voxel): how many points fall in it and how
    many of those satisfy `indicator` (e.g. distance <= 3cm). These per-block sums are
    what the bootstrap resamples, so within-block correlation is preserved."""
    if len(points_m) == 0:
        return np.array([]), np.array([]), 0
    q = np.floor(points_m / block_m).astype(np.int64)
    _, inv = np.unique(q, axis=0, return_inverse=True)
    nb = int(inv.max()) + 1
    within = np.bincount(inv, weights=indicator.astype(float), minlength=nb)
    total = np.bincount(inv, minlength=nb).astype(float)
    return within, total, nb


def bootstrap_f1_draws(acc_pts, acc_ind, comp_pts, comp_ind, block_m, B, rng):
    """B block-bootstrap draws of F1@3cm (%). Accuracy blocks (kept source points) and
    completeness blocks (target points) are resampled independently each iteration."""
    aw, at, anb = _block_parts(acc_pts, acc_ind, block_m)
    cw, ct, cnb = _block_parts(comp_pts, comp_ind, block_m)
    if anb == 0 or cnb == 0:
        return None, anb, cnb
    ai = rng.integers(0, anb, size=(B, anb))
    acc_b = aw[ai].sum(1) / at[ai].sum(1)
    ci = rng.integers(0, cnb, size=(B, cnb))
    comp_b = cw[ci].sum(1) / ct[ci].sum(1)
    denom = acc_b + comp_b
    f1_b = np.where(denom > 0, 2 * acc_b * comp_b / denom, 0.0) * 100.0
    return f1_b, anb, cnb


# --- main ------------------------------------------------------------------------

def load_reference(path) -> o3d.geometry.PointCloud:
    """Read a reference (LiDAR) cloud and drop exact-duplicate points before any
    metric touches it.

    The delivered clouds carry a lot of them - 10-26% of the points on these objects
    (bus_stop_001 41337 -> 31692, flashlight_003 22637 -> 16690, bollard_003 2818 ->
    2526) - left over where overlapping scan passes cover the same surface twice, at
    byte-identical coordinates. They are not extra information, but they do skew the
    metrics: completeness is averaged over target points, so a duplicated point is
    counted as many times as it appears, quietly weighting the score toward whatever
    the scanner happened to pass twice. Distances themselves are unaffected (a
    duplicate is its own nearest neighbour), so this only removes the double-counting.
    """
    pcd = o3d.io.read_point_cloud(str(path))
    pts = np.asarray(pcd.points)
    _, first_idx = np.unique(pts, axis=0, return_index=True)
    if len(first_idx) < len(pts):
        # select_by_index (rather than rebuilding from the array) so colors/normals survive
        pcd = pcd.select_by_index(np.sort(first_idx).tolist())
        print(f"       deduplicated reference: {len(pts)} -> {len(first_idx)} points "
              f"({100 * (1 - len(first_idx) / len(pts)):.1f}% exact duplicates removed)", flush=True)
    return pcd



def write_summary_xlsx(summary: list[dict], sensitivity: dict, path: Path) -> None:
    """Write the capture-comparison table as .xlsx.

    Generated here rather than by hand: the previous file was produced once, outside any
    script, and silently went stale - by the time it was noticed it still carried the old
    T1/T2/T3 wording AND pre-density-match numbers, months out of date. Writing it from the
    same run that builds the page means the two can no longer disagree.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    thresh = [f"{t:g}cm" for t in THRESHOLDS_CM]
    headers = ["object", "capture approach", "method", "exp_id"]
    for k in thresh:
        headers += [f"accuracy@{k} (%)", f"completeness@{k} (%)", f"F1@{k} (%)"]
    headers += ["ΔF1@10-3cm (pp)", "F1@3cm 95% CI lo", "F1@3cm 95% CI hi",
                "accuracy median (cm)", "completeness median (cm)", "inlier RMSE@3cm (cm)",
                "reg-rate (%)", "points raw", "points density-matched", "excluded as gap",
                "DBSCAN (ft/eps/mp)"]

    wb = Workbook(); ws = wb.active; ws.title = "capture_comparison"
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F3864")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    rows = sorted(summary, key=lambda r: (r["object"], r["method"], r["approach"]))
    for r in rows:
        d = OBJECT_META[r["object"]]["dbscan"]
        vals = [r["object"], APPROACH_LABEL[r["approach"]], METHOD_LABEL[r["method"]], r["exp_id"]]
        for k in thresh:
            vals += [r.get(f"acc_{k}"), r.get(f"comp_{k}"), r.get(f"f1_{k}")]
        vals += [r.get("f1_delta_10_3"), r.get("f1_ci_lo"), r.get("f1_ci_hi"),
                 r["accuracy_median_cm"], r["completeness_median_cm"], r.get("inlier_rmse_3cm"),
                 r.get("reg_rate"), r["raw_points"], r["matched_points"], r.get("n_excluded"),
                 f"{d['ft']:g}/{d['eps']:g}/{d['mp']:g}"]
        ws.append(vals)

    # bold the best F1 per (object, method) group, independently at each threshold
    groups: dict[tuple, list[int]] = {}
    for i, r in enumerate(rows):
        groups.setdefault((r["object"], r["method"]), []).append(i)
    for k_i, k in enumerate(thresh):
        col = 4 + 3 * k_i + 3
        for idxs in groups.values():
            best = max(idxs, key=lambda i: rows[i].get(f"f1_{k}") or 0)
            ws.cell(row=best + 2, column=col).font = Font(bold=True)

    top = Border(top=Side(style="medium", color="9AA07A"))
    for idxs in groups.values():
        for c in range(1, len(headers) + 1):
            ws.cell(row=idxs[0] + 2, column=c).border = top

    widths = [21, 24, 12, 9] + [14, 16, 11] * len(thresh) + [15, 15, 15, 15, 17, 17, 11, 12, 16, 14, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "E2"; ws.row_dimensions[1].height = 34

    # --- sheet 2: the significance test, so "the difference is not resolvable" can be
    # cited from the table itself rather than from a side JSON. Every pairwise F1@3cm
    # difference with its block-bootstrap 95% CI; a CI spanning 0 means the two capture
    # approaches are not distinguishable for that object/method at this sample size.
    ws2 = wb.create_sheet("significance")
    boot = sensitivity.get("bootstrap", {})
    ws2.append([f"Pairwise F1@3cm differences, {boot.get('n_draws','?')} spatial block-bootstrap draws, "
                f"{boot.get('block_cm','?')} cm blocks. CI spanning 0 = not resolvable."])
    ws2["A1"].font = Font(italic=True, color="585D54")
    ws2.append([])
    head2 = ["object", "method", "pair", "ΔF1@3cm (pp)", "95% CI lo", "95% CI hi", "resolvable?"]
    ws2.append(head2)
    for c in ws2[3]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F3864")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for e in sensitivity.get("per_object_method", []):
        for pw in e.get("pairwise", []):
            ws2.append([e["object"], METHOD_LABEL.get(e["method"], e["method"]), pw["label"],
                        pw["delta"], pw["ci_lo"], pw["ci_hi"],
                        "no - CI spans 0" if pw["includes_zero"] else "yes"])
    for row in ws2.iter_rows(min_row=4, min_col=7, max_col=7):
        for c in row:
            if c.value == "yes":
                c.font = Font(bold=True, color="0D8054")
    for col, w in zip("ABCDEFG", (21, 12, 10, 15, 12, 12, 17)):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A4"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"Wrote {path.relative_to(PROJECT_ROOT)}", flush=True)


def main() -> None:
    reg_rates = reg_rates_from_metrics()
    for _e, _v in reg_rates_from_experiments_yaml().items():
        reg_rates.setdefault(_e, _v)   # jsonl wins where present; yaml fills the gaps

    # per-object reference: load once, spacing + histogram + embedded target positions
    objects_data: list[dict] = []
    ref_cache: dict[str, tuple] = {}  # object_id -> (o3d ref pcd, spacing_m, target_pos_embed_idx?)

    for obj_id, ref_path in REFS.items():
        print(f"[ref] {obj_id}: loading {ref_path.name}", flush=True)
        ref = load_reference(ref_path)
        rpts = np.asarray(ref.points)
        spacing = nn_spacing_all(rpts)
        median_spacing = float(np.median(spacing))
        height_m = float(rpts[:, 2].max() - rpts[:, 2].min())  # vertical (Z) extent of the reference
        print(f"       {len(rpts)} pts, median NN spacing = {median_spacing*100:.3f} cm, "
              f"height = {height_m:.2f} m", flush=True)

        # histogram of NN-spacing (cm)
        spacing_cm = spacing * 100.0
        edges = np.linspace(0.0, HIST_MAX_CM, HIST_BINS + 1)
        counts, _ = np.histogram(np.clip(spacing_cm, 0, HIST_MAX_CM), bins=edges)
        overflow_pct = float(np.mean(spacing_cm > HIST_MAX_CM) * 100.0)

        # embedded target positions (shared by this object's 6 panels)
        tsub = subsample(len(rpts), TARGET_CAP)
        target_pos_embed = rpts[tsub]

        meta = OBJECT_META[obj_id]
        objects_data.append({
            "id": obj_id,
            "title": meta["title"],
            "shape": meta["shape"],
            "ref_note": meta["ref_note"],
            "ref_spacing_cm": round(median_spacing * 100, 4),
            "height_m": round(height_m, 2),
            "dbscan": meta["dbscan"],
            "hist": {
                "n": int(len(rpts)),
                "mean_cm": float(spacing_cm.mean()),
                "median_cm": float(np.median(spacing_cm)),
                "p95_cm": float(np.percentile(spacing_cm, 95)),
                "max_cm": float(spacing_cm.max()),
                "overflow_pct": overflow_pct,
                "counts": counts.astype(int).tolist(),
                "bin_edges_cm": [round(e, 4) for e in edges.tolist()],
            },
            "target_pos": b64f(target_pos_embed),
            "n_target_total": int(len(rpts)),
            "panels": [],  # filled below, in the ENTRIES order
        })
        ref_cache[obj_id] = (ref, rpts, median_spacing)

    objects_by_id = {o["id"]: o for o in objects_data}

    # per reconstruction
    panels: dict[str, dict] = {}
    boot_rng = np.random.default_rng(BOOT_SEED)
    f1_draws: dict[str, np.ndarray] = {}  # key -> B bootstrap F1@3cm draws, for the sensitivity CIs
    for obj_id, method, approach, exp_id, rel in ENTRIES:
        key = f"{obj_id}__{method}__{approach}"
        ref, rpts, spacing_m = ref_cache[obj_id]
        src_path = PROJECT_ROOT / rel
        print(f"[panel] {key}  ({exp_id})  <- {src_path.name}", flush=True)
        src = o3d.io.read_point_cloud(str(src_path))
        raw_points = len(src.points)

        matched = src.voxel_down_sample(VOXEL_M)
        mpts = np.asarray(matched.points)
        n_matched = len(mpts)

        # accuracy direction (source -> target) on density-matched cloud
        d_s2t = np.asarray(matched.compute_point_cloud_distance(ref))  # meters
        d_s2t_cm = d_s2t * 100.0
        # completeness direction (target -> source), against FULL matched cloud
        d_t2s_cm = np.asarray(ref.compute_point_cloud_distance(matched)) * 100.0

        below_mask = d_s2t_cm <= FLOOR_CM
        cand_mask = ~below_mask
        n_below_true = int(below_mask.sum())
        n_cand_true = int(cand_mask.sum())

        below_idx = np.where(below_mask)[0]
        cand_idx = np.where(cand_mask)[0]
        below_sel = below_idx[subsample(len(below_idx), EMBED_CAP)]
        cand_sel = cand_idx[subsample(len(cand_idx), EMBED_CAP)]

        # default-params sanity metrics (DBSCAN gap-exclusion, then Acc/Comp/F1@3cm)
        dcfg = OBJECT_META[obj_id]["dbscan"]
        far_m = dcfg["ft"] / 100.0
        cand_far_idx = np.where(d_s2t > far_m)[0]
        gap_mask = np.zeros(n_matched, dtype=bool)
        if len(cand_far_idx) > 0:
            far_pcd = o3d.geometry.PointCloud()
            far_pcd.points = o3d.utility.Vector3dVector(mpts[cand_far_idx])
            labels = np.array(far_pcd.cluster_dbscan(eps=dcfg["eps"] / 100.0, min_points=dcfg["mp"]))
            gap_mask[cand_far_idx[labels >= 0]] = True
        kept = ~gap_mask
        n_excluded = int(gap_mask.sum())
        d_kept_cm = d_s2t_cm[kept]
        by_threshold = {}
        for _t in THRESHOLDS_CM:
            _a = float(np.mean(d_kept_cm <= _t)) if d_kept_cm.size else 0.0
            _c = float(np.mean(d_t2s_cm <= _t)) if d_t2s_cm.size else 0.0
            by_threshold[f"{_t:g}cm"] = (_a, _c, f_score(_a, _c))
        acc_3, comp_3 = by_threshold["3cm"][0], by_threshold["3cm"][1]
        inliers_3 = d_kept_cm[d_kept_cm <= 3.0]
        inlier_rmse_3 = float(np.sqrt(np.mean(inliers_3 ** 2))) if inliers_3.size else float("nan")
        f1_3 = f_score(acc_3, comp_3)
        acc_median = float(np.median(d_kept_cm)) if d_kept_cm.size else float("nan")
        comp_median = float(np.median(d_t2s_cm)) if d_t2s_cm.size else float("nan")

        # 95% spatial block-bootstrap CI on F1@3cm (see B_BOOT/BLOCK_CM notes above)
        f1_b, n_blocks_acc, n_blocks_comp = bootstrap_f1_draws(
            mpts[kept], d_kept_cm <= 3.0, rpts, d_t2s_cm <= 3.0, BLOCK_CM / 100.0, B_BOOT, boot_rng,
        )
        if f1_b is not None:
            f1_draws[key] = f1_b
            f1_ci_lo, f1_ci_hi = (float(x) for x in np.percentile(f1_b, [2.5, 97.5]))
        else:
            f1_ci_lo = f1_ci_hi = float("nan")

        print(f"          matched={n_matched}  excluded={n_excluded}  "
              f"F1@3={f1_3*100:.1f}% [{f1_ci_lo:.1f},{f1_ci_hi:.1f}]  acc_med={acc_median:.2f}cm  "
              f"comp_med={comp_median:.2f}cm  inlierRMSE@3={inlier_rmse_3:.2f}cm  "
              f"blocks(acc/comp)={n_blocks_acc}/{n_blocks_comp}", flush=True)

        panels[key] = {
            "object": obj_id,
            "method": method,
            "approach": approach,
            "exp_id": exp_id,
            "label": f"{METHOD_LABEL[method]} · T{approach}",
            "raw_points": raw_points,
            "matched_points": n_matched,
            "reg_rate": reg_rates.get(exp_id),
            "accuracy_median_cm": round(acc_median, 3),
            "completeness_median_cm": round(comp_median, 3),
            "below_pos": b64f(mpts[below_sel]),
            "below_dist_cm": b64f(d_s2t_cm[below_sel]),
            "candidate_pos": b64f(mpts[cand_sel]),
            "candidate_dist_cm": b64f(d_s2t_cm[cand_sel]),
            "target_dist_cm": b64f(d_t2s_cm),
            "n_source_total": n_matched,
            "n_below_true": n_below_true,
            "n_candidates_true": n_cand_true,
            "n_target_total": int(len(rpts)),
            "default": {
                "f1_3cm": round(f1_3 * 100, 2),
                "acc_3cm": round(acc_3 * 100, 2),
                "comp_3cm": round(comp_3 * 100, 2),
                **{f"{m}_{k}": round(v * 100, 2)
                   for k, (a, c, f) in by_threshold.items()
                   for m, v in (("acc", a), ("comp", c), ("f1", f))
                   if k != "3cm"},
                "f1_delta_10_3": round((by_threshold["10cm"][2] - by_threshold["3cm"][2]) * 100, 2),
                "inlier_rmse_3cm": round(inlier_rmse_3, 3),
                "n_excluded": n_excluded,
                "f1_ci_lo": round(f1_ci_lo, 2),
                "f1_ci_hi": round(f1_ci_hi, 2),
                "n_blocks_acc": int(n_blocks_acc),
                "n_blocks_comp": int(n_blocks_comp),
            },
        }
        objects_by_id[obj_id]["panels"].append(key)

    # ----- capture-approach sensitivity: how much does the approach move F1, per
    # (object, method), relative to the bootstrap noise? spread = max-min F1 across the
    # three approaches; its CI comes from the joint block-bootstrap draws. We also flag
    # whether the best and worst approach's F1 CIs overlap (if they don't, the capture
    # approach makes a statistically resolvable difference).
    def ci_overlap(a, b):
        return not (a[1] < b[0] or b[1] < a[0])

    sensitivity = []
    for obj in objects_data:
        methods = []
        for key in obj["panels"]:
            m = panels[key]["method"]
            if m not in methods:
                methods.append(m)
        for method in methods:
            keys = [f"{obj['id']}__{method}__{a}" for a in (1, 2, 3)]
            if not all(k in f1_draws for k in keys):
                continue
            draws = np.vstack([f1_draws[k] for k in keys])  # 3 x B
            point = np.array([panels[k]["default"]["f1_3cm"] for k in keys])
            spread_b = draws.max(0) - draws.min(0)
            best_i, worst_i = int(point.argmax()), int(point.argmin())
            best_ci = (panels[keys[best_i]]["default"]["f1_ci_lo"], panels[keys[best_i]]["default"]["f1_ci_hi"])
            worst_ci = (panels[keys[worst_i]]["default"]["f1_ci_lo"], panels[keys[worst_i]]["default"]["f1_ci_hi"])

            # PROPER significance test: bootstrap CI of the DIFFERENCE between the best and
            # worst approach (the two clouds are independent, so their draws are independent).
            # If this CI includes 0 the difference is not resolvable - this is the correct
            # test, not the (too-conservative) "do the two individual CIs overlap" eyeball.
            diff_b = draws[best_i] - draws[worst_i]
            diff_lo, diff_hi = (float(x) for x in np.percentile(diff_b, [2.5, 97.5]))
            diff_includes_zero = bool(diff_lo <= 0.0 <= diff_hi)

            # all three pairwise differences (T1-T2, T1-T3, T2-T3) with their CIs, so it's
            # visible WHICH approaches differ, not just the extreme pair.
            pairwise = []
            for i, j, lbl in [(0, 1, "T1−T2"), (0, 2, "T1−T3"), (1, 2, "T2−T3")]:
                db = draws[i] - draws[j]
                lo, hi = (float(x) for x in np.percentile(db, [2.5, 97.5]))
                pairwise.append({
                    "label": lbl, "delta": round(float(point[i] - point[j]), 2),
                    "ci_lo": round(lo, 2), "ci_hi": round(hi, 2),
                    "includes_zero": bool(lo <= 0.0 <= hi),
                })

            # histograms of the three pairwise DIFFERENCE distributions (shared bin edges),
            # order matches `pairwise`: T1-T2, T1-T3, T2-T3. A distribution's position
            # relative to 0 is the significance read.
            diffs3 = [draws[0] - draws[1], draws[0] - draws[2], draws[1] - draws[2]]
            alld = np.concatenate(diffs3)
            lo, hi = float(alld.min()), float(alld.max())
            pad = (hi - lo) * 0.04 + 1e-6
            edges = np.linspace(lo - pad, hi + pad, 37)
            hist_counts = [np.histogram(d, bins=edges)[0].astype(int).tolist() for d in diffs3]

            sensitivity.append({
                "object": obj["id"], "method": method,
                "f1": [round(float(x), 2) for x in point],
                "spread": round(float(point.max() - point.min()), 2),
                "spread_ci_lo": round(float(np.percentile(spread_b, 2.5)), 2),
                "spread_ci_hi": round(float(np.percentile(spread_b, 97.5)), 2),
                "best_approach": best_i + 1, "worst_approach": worst_i + 1,
                "best_f1": round(float(point[best_i]), 2), "worst_f1": round(float(point[worst_i]), 2),
                "best_worst_ci_overlap": ci_overlap(best_ci, worst_ci),
                "diff_ci_lo": round(diff_lo, 2), "diff_ci_hi": round(diff_hi, 2),
                "diff_includes_zero": diff_includes_zero,
                "pairwise": pairwise,
                "hist": {"edges": [round(float(e), 3) for e in edges.tolist()], "counts": hist_counts},
            })
    # per-object mean spread across its methods (headline "sensitivity of this object")
    object_sensitivity = []
    for obj in objects_data:
        rows = [s for s in sensitivity if s["object"] == obj["id"]]
        if rows:
            object_sensitivity.append({
                "object": obj["id"],
                "mean_spread": round(float(np.mean([r["spread"] for r in rows])), 2),
                "max_spread": round(float(np.max([r["spread"] for r in rows])), 2),
                "any_resolvable": any(not r["diff_includes_zero"] for r in rows),
            })
    print("\nCapture-approach sensitivity (F1@3cm best-vs-worst difference, bootstrap CI):", flush=True)
    for s in sensitivity:
        print(f"  {s['object']:22s} {s['method']:10s} spread={s['spread']:5.1f}  "
              f"diff(T{s['best_approach']}-T{s['worst_approach']}) 95% CI=[{s['diff_ci_lo']:.1f},{s['diff_ci_hi']:.1f}]  "
              f"includes_0={s['diff_includes_zero']}", flush=True)

    data = {
        "floor_cm": FLOOR_CM,
        "objects": objects_data,
        "panels": panels,
        "method_label": METHOD_LABEL,
        "approach_label": APPROACH_LABEL,
        "sensitivity": sensitivity,
        "object_sensitivity": object_sensitivity,
        "bootstrap": {"n_draws": B_BOOT, "block_cm": BLOCK_CM},
    }

    html = build_html(data)
    OUT_HTML.parent.mkdir(parents=True, exist_ok=True)
    OUT_HTML.write_text(html, encoding="utf-8")
    size_mb = OUT_HTML.stat().st_size / (1024 * 1024)
    print(f"\nWrote {OUT_HTML.relative_to(PROJECT_ROOT)}  ({size_mb:.2f} MB)", flush=True)

    # also dump the numeric summary as JSON for the optional xlsx / quick inspection
    summary_path = PROJECT_ROOT / "docs" / "tables" / "capture_comparison_summary.json"
    summary = [
        {
            "object": p["object"], "method": p["method"], "approach": p["approach"],
            "exp_id": p["exp_id"], "reg_rate": p["reg_rate"],
            "raw_points": p["raw_points"], "matched_points": p["matched_points"],
            "accuracy_median_cm": p["accuracy_median_cm"],
            "completeness_median_cm": p["completeness_median_cm"],
            **p["default"],
        }
        for p in panels.values()
    ]
    summary_path.write_text(json.dumps(summary, indent=2))
    print(f"Wrote {summary_path.relative_to(PROJECT_ROOT)}", flush=True)

    # one payload for both outputs, so the xlsx's significance sheet and the JSON can
    # never quote different intervals
    sens_payload = {"bootstrap": data["bootstrap"], "per_object_method": sensitivity,
                    "per_object": object_sensitivity}

    write_summary_xlsx(summary, sens_payload, PROJECT_ROOT / "docs" / "tables" / "capture_comparison_summary.xlsx")

    sens_path = PROJECT_ROOT / "docs" / "tables" / "capture_comparison_sensitivity.json"
    sens_path.write_text(json.dumps(sens_payload, indent=2))
    print(f"Wrote {sens_path.relative_to(PROJECT_ROOT)}", flush=True)


def build_html(data: dict) -> str:
    payload = json.dumps(data).replace("</", "<\\/")
    return HTML_HEAD + f'\n<script type="application/json" id="page-data">{payload}</script>\n' + MAIN_JS + HTML_TAIL


# HTML_HEAD / MAIN_JS / HTML_TAIL are defined in the template module below to keep this
# file readable; they are imported at module load.
from _capture_page_template import HTML_HEAD, MAIN_JS, HTML_TAIL  # noqa: E402


if __name__ == "__main__":
    sys.exit(main())
