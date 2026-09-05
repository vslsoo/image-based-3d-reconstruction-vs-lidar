"""Build site/results.html - the main result of the dissertation on one page.

Six objects x four methods, Accuracy / Completeness / F1 at 3, 5 and 10 cm against the LiDAR
reference. The site had three study pages and six per-object pages, but nowhere to see the
headline table: those numbers existed only as prose on the index cards and scattered one
object at a time across the object pages.

Nothing is recomputed here. The rows are read straight out of
docs/tables/summary_all_objects_accuracy_f1_EN.xlsx (written by
build_accuracy_f1_summary_table.py, which computes them over the full clouds), so this page
cannot drift from the workbook the thesis cites - and it needs no open3d, so it rebuilds in
under a second.

Usage:
    python src/registration/build_results_page.py
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _site_nav import NAV_CSS, nav_html  # noqa: E402

PROJECT_ROOT = Path(__file__).resolve().parents[2]
SRC_XLSX = PROJECT_ROOT / "docs" / "tables" / "summary_all_objects_accuracy_f1_EN.xlsx"
OUT_HTML = PROJECT_ROOT / "site" / "results.html"

THRESHOLDS = ["3cm", "5cm", "10cm"]

# the labels the rest of the site uses; the workbook stores the internal ids
METHOD_LABEL = {"colmap": "COLMAP", "hloc_colmap": "hloc + COLMAP",
                "mast3r_ga": "MASt3R-GA", "vggt": "VGGT"}

# object_id -> the page it has on this site, and the name the index uses for it
OBJECT_PAGE = {
    "bus_stop_002": ("bus_stop.html", "bus shelter"),
    "information_sign_002": ("information_sign.html", "information sign"),
    "bench_004": ("bench.html", "bench"),
    "bollard_003": ("bollard.html", "bollard"),
    "flashlight_004": ("flashlight.html", "lamppost"),
    "bus_stop_sign_002": ("bus_stop_sign.html", "bus-stop sign"),
}


def read_rows() -> list[dict]:
    ws = load_workbook(SRC_XLSX)["summary"]
    hdr = [c.value for c in ws[1]]
    rows, current = [], None
    for raw in ws.iter_rows(min_row=2, values_only=True):
        r = dict(zip(hdr, raw))
        # the workbook writes the object id, shape and size once per group, on its first row
        if r["object_id"]:
            current = {k: r[k] for k in ("object_id", "shape", "ref. spacing (cm)", "length (cm)",
                                         "width (cm)", "height (cm)", "DBSCAN mode", "reference note")}
        row = {**current, **{k: v for k, v in r.items() if v is not None or k not in current}}
        row["object_id"] = current["object_id"]
        rows.append(row)
    return rows


def page_data(rows: list[dict]) -> dict:
    objects: dict[str, dict] = {}
    for r in rows:
        obj = objects.setdefault(r["object_id"], {
            "id": r["object_id"],
            "page": OBJECT_PAGE.get(r["object_id"], ("", r["object_id"]))[0],
            "name": OBJECT_PAGE.get(r["object_id"], ("", r["object_id"]))[1],
            "shape": r["shape"],
            "size_cm": [r["length (cm)"], r["width (cm)"], r["height (cm)"]],
            "dbscan_mode": r["DBSCAN mode"],
            "ref_note": r["reference note"],
            "methods": [],
        })
        obj["methods"].append({
            "method": METHOD_LABEL.get(r["method"], r["method"]),
            "acc_median_cm": r["accuracy median (cm)"],
            "comp_median_cm": r["completeness median (cm)"],
            "rmse_mm": r["alignment RMSE (mm)"],
            "raw_points": r["raw points"],
            "matched_points": r["matched points (1cm voxel)"],
            "delta_10_3": r["ΔF1@10-3cm (pp)"],
            **{f"{m}_{t}": r[f"{name}@{t} (%)"]
               for t in THRESHOLDS
               for m, name in (("acc", "accuracy"), ("comp", "completeness"), ("f1", "F1"))},
        })
    return {"objects": list(objects.values()), "thresholds": THRESHOLDS}


HTML = """<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Main results — 6 objects × 4 methods vs LiDAR</title>
</head>
<body>
<style>
  :root {
    --bg:#ffffff; --panel:#ffffff; --panel-border:#d7d4c8; --text:#181a17; --text-dim:#585d54; --text-faint:#8b9084;
    --accent:#17805f; --accent-soft:#d9ece3; --code-bg:#f5f4ef; --best:#0d8054; --best-soft:#e3f1ea; --red:#e16b3e;
  }
  @media (prefers-color-scheme: dark) {
    :root { --bg:#ffffff; --panel:#ffffff; --panel-border:#d7d4c8; --text:#181a17; --text-dim:#585d54;
            --text-faint:#8b9084; --accent:#17805f; --accent-soft:#d9ece3; --code-bg:#f5f4ef; --best:#0d8054;
            --best-soft:#e3f1ea; --red:#e16b3e; }
  }
  * { box-sizing:border-box; }
  body { margin:0; background:var(--bg); color:var(--text); line-height:1.45;
         font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Helvetica,Arial,sans-serif; }
  .page { max-width:1560px; margin:0 auto; padding:28px 24px 72px; display:flex; flex-direction:column; gap:22px; }
  a { color:var(--accent); }
  .eyebrow { font-size:11.5px; font-weight:600; letter-spacing:.09em; text-transform:uppercase; color:var(--accent); }
  h1 { font-size:23px; font-weight:650; margin:4px 0 2px; letter-spacing:-.01em; }
  h2 { font-size:19px; font-weight:650; margin:0 0 2px; }
  .subtitle { color:var(--text-dim); font-size:13.5px; max-width:92ch; }
  b, strong { color:var(--text); font-weight:600; }
  .mono { font-family:ui-monospace,"SF Mono","Cascadia Code",Menlo,Consolas,monospace; font-variant-numeric:tabular-nums; }
  section { display:flex; flex-direction:column; gap:12px; }
  hr.sep { border:none; border-top:1px solid var(--panel-border); margin:2px 0; }
  .tabs { display:flex; flex-wrap:wrap; gap:5px; }
  .tab-btn { font-family:inherit; font-size:10.5px; padding:3px 9px; border-radius:6px; border:1px solid var(--panel-border);
             background:transparent; color:var(--text-dim); cursor:pointer; }
  .tab-btn:hover { border-color:var(--accent); color:var(--text); }
  .tab-btn.active { background:var(--accent-soft); border-color:var(--accent); color:var(--text); font-weight:600; }
  .grid-wrap { overflow-x:auto; }
  table.summary { border-collapse:collapse; font-size:12px; min-width:1080px; }
  table.summary th, table.summary td { padding:6px 10px; border-bottom:1px solid var(--panel-border); text-align:right; white-space:nowrap; }
  table.summary th { font-weight:650; color:var(--text-dim); position:sticky; top:0; background:var(--panel); }
  table.summary td.txt, table.summary th.txt { text-align:left; }
  table.summary td.f1cell { font-weight:700; font-variant-numeric:tabular-nums; }
  table.summary tr.best td.f1cell { color:var(--best); }
  table.summary td.best-cell { background:var(--best-soft); border-radius:4px; }
  table.summary tbody tr:hover { background:color-mix(in srgb, var(--accent-soft) 40%, transparent); }
  .grouprule td { border-top:2px solid var(--panel-border); }
  .note { font-size:11.5px; color:var(--text-faint); }
  .bar { display:inline-block; height:7px; border-radius:3px; background:var(--accent); opacity:.75; vertical-align:1px; }
__NAV_CSS__
</style>

<div class="page">
  __SITE_NAV__
  <div>
    <div class="eyebrow">Main results · gap-aware Chamfer vs the LiDAR reference</div>
    <h1>Six objects, four methods, one table</h1>
    <div class="subtitle">
      Every reconstruction on this site, scored the same way: density-matched onto the reference's
      1&nbsp;cm grid, source↔target Chamfer distances, gap-aware Accuracy / Completeness / F1. Each object
      uses the gap-detection setting its own page defaults to, and the numbers are computed over the full
      clouds — identical to <span class="mono">docs/tables/summary_all_objects_accuracy_f1.xlsx</span> and to
      the “Main results” sheet of <span class="mono">FINAL_results.xlsx</span>.
    </div>
  </div>

  <hr class="sep">
  <section>
    <div style="display:flex; align-items:baseline; gap:14px; flex-wrap:wrap;">
      <h2 id="table-title">Accuracy, Completeness and F1 at 3&nbsp;cm</h2>
      <div class="tabs" id="thr-toggle">
        <button class="tab-btn active" data-thr="3cm">3 cm</button>
        <button class="tab-btn" data-thr="5cm">5 cm</button>
        <button class="tab-btn" data-thr="10cm">10 cm</button>
      </div>
    </div>
    <div class="subtitle">
      Best F1 per object is highlighted. <b>ΔF1@10−3</b> is how much F1 recovers when the threshold widens
      from 3 to 10&nbsp;cm: small means the surface is already where it should be, large means it is there
      but displaced — and a row that stays low even at 10&nbsp;cm never reconstructed that geometry at all.
    </div>
    <div class="grid-wrap"><div id="table-wrap"></div></div>
    <div class="note" id="table-note"></div>
  </section>
</div>

<script type="application/json" id="page-data">__PAYLOAD__</script>
<script>
const DATA = JSON.parse(document.getElementById('page-data').textContent);
let thr = '3cm';

function fmt(v, d = 1) { return v == null ? '—' : (+v).toFixed(d); }

function buildTable() {
  let h = '<table class="summary"><thead><tr>'
    + '<th class="txt">Object</th><th class="txt">Method</th>'
    + `<th id="th-f1">F1@${thr}</th><th></th><th id="th-acc">Acc@${thr}</th><th id="th-comp">Comp@${thr}</th>`
    + '<th>ΔF1@10−3</th><th>Acc median (cm)</th><th>Comp median (cm)</th><th>align. RMSE (mm)</th>'
    + '<th>#pts (raw→matched)</th></tr></thead><tbody>';
  for (const obj of DATA.objects) {
    const best = Math.max(...obj.methods.map(m => m[`f1_${thr}`] ?? 0));
    obj.methods.forEach((m, i) => {
      const isBest = (m[`f1_${thr}`] ?? 0) === best;
      const f1 = m[`f1_${thr}`];
      h += `<tr${i === 0 ? ' class="grouprule"' : ''}${isBest ? ' style="font-weight:500"' : ''}>`
        + `<td class="txt">${i === 0 ? `<a href="${obj.page}">${obj.name}</a>`
             + `<div class="note">${obj.size_cm.map(v => (v / 100).toFixed(2)).join(' × ')} m · ${obj.dbscan_mode}</div>` : ''}</td>`
        + `<td class="txt">${m.method}</td>`
        + `<td class="f1cell${isBest ? ' best-cell' : ''}">${fmt(f1)}</td>`
        + `<td style="width:110px; text-align:left;"><span class="bar" style="width:${Math.max(0, (f1 ?? 0)) }px"></span></td>`
        + `<td>${fmt(m[`acc_${thr}`])}</td><td>${fmt(m[`comp_${thr}`])}</td>`
        + `<td>${fmt(m.delta_10_3)}</td>`
        + `<td>${fmt(m.acc_median_cm, 2)}</td><td>${fmt(m.comp_median_cm, 2)}</td><td>${fmt(m.rmse_mm)}</td>`
        + `<td class="mono">${(m.raw_points ?? 0).toLocaleString('en-US')}→${(m.matched_points ?? 0).toLocaleString('en-US')}</td>`
        + '</tr>';
    });
  }
  h += '</tbody></table>';
  document.getElementById('table-wrap').innerHTML = h;
  document.getElementById('table-title').innerHTML =
    `Accuracy, Completeness and F1 at ${thr.replace('cm', '&nbsp;cm')}`;

  // one line of context that follows the threshold, rather than a paragraph that cannot
  const rows = DATA.objects.flatMap(o => o.methods.map(m => ({ obj: o, m })));
  const top = rows.reduce((a, b) => (b.m[`f1_${thr}`] ?? 0) > (a.m[`f1_${thr}`] ?? 0) ? b : a);
  const bottom = rows.reduce((a, b) => (b.m[`f1_${thr}`] ?? 0) < (a.m[`f1_${thr}`] ?? 0) ? b : a);
  document.getElementById('table-note').innerHTML =
    `At ${thr}: best is <b>${top.obj.name} · ${top.m.method}</b> at ${fmt(top.m[`f1_${thr}`])}%, `
    + `worst is <b>${bottom.obj.name} · ${bottom.m.method}</b> at ${fmt(bottom.m[`f1_${thr}`])}%. `
    + `Four of the six references are incomplete (parts were never scanned); those gaps are excluded by `
    + `DBSCAN at each object's own setting, shown under its name — see the `
    + `<a href="tuner.html">gap tuner</a> for what that line depends on.`;
}

document.querySelectorAll('#thr-toggle .tab-btn').forEach(b => b.addEventListener('click', () => {
  document.querySelectorAll('#thr-toggle .tab-btn').forEach(x => x.classList.remove('active'));
  b.classList.add('active');
  thr = b.dataset.thr;
  buildTable();
}));

buildTable();
</script>
</body>
</html>
"""


def main() -> None:
    rows = read_rows()
    data = page_data(rows)
    html = (HTML.replace("__NAV_CSS__", NAV_CSS)
                .replace("__SITE_NAV__", nav_html("results"))
                .replace("__PAYLOAD__", json.dumps(data).replace("</", "<\\/")))
    OUT_HTML.write_text(html, encoding="utf-8")
    n = sum(len(o["methods"]) for o in data["objects"])
    print(f"Wrote {OUT_HTML.relative_to(PROJECT_ROOT)} "
          f"({len(data['objects'])} objects, {n} rows, {OUT_HTML.stat().st_size / 1024:.0f} KB)")


if __name__ == "__main__":
    sys.exit(main())
