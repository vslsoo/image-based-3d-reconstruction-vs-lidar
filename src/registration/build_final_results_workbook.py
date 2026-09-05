"""Assemble docs/tables/FINAL_results.xlsx - one workbook holding every result the
dissertation reports, so the thesis can cite a single file instead of five.

It does not recompute anything. Each sheet is copied verbatim from the table its own
builder already wrote, so this file can never disagree with the site:

  Main results        <- summary_all_objects_accuracy_f1_EN.xlsx   (build_accuracy_f1_summary_table.py)
  Capture strategy    <- capture_comparison_summary.xlsx           (build_capture_comparison_page.py)
  Capture significance<- capture_comparison_summary.xlsx sheet 2
  Frame count         <- frame_count_study_summary.xlsx            (build_frame_count_study_page.py)
  Frame significance  <- frame_count_study_summary.xlsx sheet 2
  Compute cost        <- performance_study_summary.xlsx            (build_performance_study_page.py)
  Experiment index    <- config/experiments.yaml (every exp_id cited above, with its
                         object, method, image count and registration rate)

Regenerate after any of those builders runs:
    python src/registration/build_final_results_workbook.py
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[2]
TABLES = PROJECT_ROOT / "docs" / "tables"
OUT = TABLES / "FINAL_results.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF")

# (destination sheet, source workbook, source sheet, one-line note placed above the table)
SOURCES = [
    ("Main results", "summary_all_objects_accuracy_f1_EN.xlsx", "summary",
     "6 objects x 4 methods. Accuracy / completeness / F1 at 3, 5 and 10 cm against the LiDAR "
     "reference, on a shared 1 cm grid. Alignment RMSE is measured from the same aligned clouds."),
    ("Capture strategy", "capture_comparison_summary.xlsx", "capture_comparison",
     "2 objects x 3 capture approaches x 2 methods. T1 = close-range + distant, T2 = close-range "
     "only, T3 = distant only."),
    ("Capture significance", "capture_comparison_summary.xlsx", "significance",
     "Pairwise F1@3cm differences between capture approaches, block bootstrap. "
     "A CI spanning 0 means the approaches are not distinguishable."),
    ("Frame count", "frame_count_study_summary.xlsx", "frame_count_study",
     "2 objects x 4 frame counts x method. Nested subsets: each larger set contains the smaller "
     "ones, so only the frame count changes."),
    ("Frame significance", "frame_count_study_summary.xlsx", "significance",
     "Paired differences between frame counts, for F1 and accuracy. Paired: the same resampled "
     "blocks feed both sides, so block-to-block variation cancels."),
    ("Compute cost", "performance_study_summary.xlsx", "performance_vs_N",
     "Wall-clock time and peak RAM/VRAM vs frame count, all runs on one NVIDIA L40S. "
     "Note the methods work at different resolutions - COLMAP 3200 px, hloc 1024 px, "
     "both feed-forward models ~512 px."),
]


def style_header(ws, row: int) -> None:
    for cell in ws[row]:
        if cell.value is None:
            continue
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize(ws, first_data_row: int) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows(min_row=first_data_row):
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(max(w + 2, 9), 30)


def copy_sheet(wb: Workbook, dest_name: str, src_file: str, src_sheet: str, note: str) -> int:
    src_path = TABLES / src_file
    if not src_path.exists():
        print(f"  SKIPPED {dest_name}: {src_file} not found")
        return 0
    src = load_workbook(src_path, data_only=True)
    if src_sheet not in src.sheetnames:
        print(f"  SKIPPED {dest_name}: no sheet {src_sheet!r} in {src_file}")
        return 0
    s = src[src_sheet]
    ws = wb.create_sheet(dest_name[:31])

    ws.append([note])
    ws["A1"].font = Font(italic=True, color="585D54")
    ws.append([f"source: docs/tables/{src_file} · sheet {src_sheet}"])
    ws["A2"].font = Font(italic=True, size=9, color="8B9084")
    ws.append([])

    rows = list(s.iter_rows(values_only=True))
    # some source sheets carry their own caption line before the header; find the header by
    # taking the first row where every leading cell is filled
    header_i = next((i for i, r in enumerate(rows)
                     if r and r[0] is not None and sum(v is not None for v in r) > 2), 0)
    for r in rows[header_i:]:
        # the performance sheet carries the internal "_test_1" capture suffix that every
        # other sheet drops; strip it so one object reads the same name throughout
        ws.append([v.replace("_test_1", "") if isinstance(v, str) and "_test_1" in v else v
                   for v in r])
    style_header(ws, 4)
    ws.freeze_panes = "A5"
    autosize(ws, 4)
    n = len(rows) - header_i - 1
    print(f"  {dest_name:<22}{n:>4} rows   <- {src_file} [{src_sheet}]")
    return n


def experiment_index(wb: Workbook, cited: set[str]) -> None:
    """Every exp_id used above, resolved against config/experiments.yaml."""
    text = (PROJECT_ROOT / "config" / "experiments.yaml").read_text()
    ws = wb.create_sheet("Experiment index")
    ws.append(["Every experiment cited in the sheets above, resolved against config/experiments.yaml."])
    ws["A1"].font = Font(italic=True, color="585D54")
    ws.append([])
    ws.append(["exp_id", "date", "object_id", "method", "images used", "registered", "reg-rate (%)"])
    style_header(ws, 3)

    found = 0
    for m in re.finditer(r"^  (exp_\d+):\n(.*?)(?=^  exp_|\Z)", text, re.S | re.M):
        exp_id, body = m.group(1), m.group(2)
        if exp_id not in cited:
            continue
        date = re.search(r"date: (\S+)", body)
        obj = re.search(r"object_id: (\S+)", body)
        meth = re.search(r"method: (\S+)", body)
        reg = re.search(r"Registered images: (\d+)/(\d+)", body)
        ws.append([
            exp_id,
            date.group(1) if date else None,
            obj.group(1) if obj else None,
            meth.group(1) if meth else None,
            int(reg.group(2)) if reg else None,
            int(reg.group(1)) if reg else None,
            round(int(reg.group(1)) / int(reg.group(2)) * 100, 1) if reg and int(reg.group(2)) else None,
        ])
        found += 1
    ws.freeze_panes = "A4"
    autosize(ws, 3)
    print(f"  {'Experiment index':<22}{found:>4} rows   <- config/experiments.yaml")


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)

    print(f"Building {OUT.relative_to(PROJECT_ROOT)}")
    for dest, src_file, src_sheet, note in SOURCES:
        copy_sheet(wb, dest, src_file, src_sheet, note)

    # collect every exp_id mentioned anywhere in the copied sheets
    cited: set[str] = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if isinstance(v, str):
                    cited.update(re.findall(r"exp_\d+", v))
    experiment_index(wb, cited)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"\nWrote {OUT.relative_to(PROJECT_ROOT)}  ({OUT.stat().st_size / 1024:.0f} KB, "
          f"{len(wb.sheetnames)} sheets)")


if __name__ == "__main__":
    main()
