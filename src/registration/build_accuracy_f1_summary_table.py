"""Regenerate docs/tables/summary_all_objects_accuracy_f1.xlsx from the CURRENT contents of
config/objects.yaml's site (build_object_page.py's MERGED_OBJECTS) - one row per (object,
method), matching exactly what's live on each per-object report page at its OWN default
DBSCAN gap-tuner setting (or "no DBSCAN" for the two objects whose honest default is the
Ignore-DBSCAN checkbox).

The previous version of this table was stale: it mixed some now-removed far-view-only
captures (bus_stop_001, bench_003, flashlight_003, bus_stop_sign_001) with what are now the
ONLY captures for other objects (bollard_003, information_sign_002) - a leftover from before
those objects had a second capture. This rebuild reads MERGED_OBJECTS directly (imported from
build_object_page.py, not re-declared) so it can never drift from what the live pages show:
6 physical objects x 4 methods = 24 rows, one row per (object, method) using each object's
single remaining capture.

Numbers are computed exactly (no subsampling) - the live pages' WebGL viewer subsamples for
render performance and population-corrects the estimate; this script computes directly over
the full aligned cloud, so these are the precise values, not an approximation of them.

Usage:
    python -u src/registration/build_accuracy_f1_summary_table.py
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import numpy as np
import open3d as o3d

sys.path.insert(0, str(Path(__file__).resolve().parent))
from build_object_page import MERGED_OBJECTS, METHOD_ORDER, DEFAULT_DBSCAN_SLIDERS  # noqa: E402

PROJECT_ROOT = Path(__file__).resolve().parents[2]
OUT_XLSX = PROJECT_ROOT / "docs" / "tables" / "summary_all_objects_accuracy_f1.xlsx"
OUT_XLSX_EN = PROJECT_ROOT / "docs" / "tables" / "summary_all_objects_accuracy_f1_EN.xlsx"
OUT_JSON = PROJECT_ROOT / "docs" / "tables" / "summary_all_objects_accuracy_f1.json"
# Same rows as the workbook, machine-readable, keyed the way the per-object pages key their
# panels ("<capture_id>__<method>"). build_object_page.py --relayout injects it into each
# page so that, while the tuner sits at that page's defaults, the panels show these exact
# numbers instead of the browser's subsample estimate of them. See that script's --relayout.

VOXEL_M = 0.01
# Voxel size used to density-match every reconstruction to the reference, in metres.
# Fixed at the 1 cm grid the LiDAR references are actually delivered on, rather than
# derived per-object from the reference's median NN-spacing. Two reasons: the delivered
# clouds are already thinned onto that grid, so the grid pitch is the real sampling
# limit; and the measured median NN-spacing sits *below* the pitch because overlapping
# scan passes leave exact-duplicate points at zero distance (bus_stop_001 reads 1.00 cm
# as delivered, 1.41 cm once duplicates are dropped). A fixed voxel also keeps the
# downsampling identical across objects and methods.

FLOOR_CM = 3.0
THRESHOLDS_CM = [3.0, 5.0, 10.0]
# Accuracy/completeness/F1 are reported at all three thresholds now, not just 3cm - 3cm is
# still what the DBSCAN gap-mask itself is built against (FLOOR_CM/the live tuner's floor),
# so it stays the "primary" bolded-winner column; 5cm/10cm are additional, looser cuts over
# the exact same kept/candidate split (only the <= comparison changes, not the gap mask).
METHOD_LABEL_RU = {"mast3r_ga": "mast3r_ga", "vggt": "vggt", "colmap": "colmap", "hloc_colmap": "hloc_colmap"}

SHAPE_RU = {
    "bus_stop": "павильон", "information_sign": "стойка", "bench": "скамья",
    "bollard": "столбик", "flashlight": "фонарь ~6м", "bus_stop_sign": "знак ~3.7м",
}
SHAPE_EN = {
    "bus_stop": "bus shelter", "information_sign": "information sign", "bench": "bench",
    "bollard": "bollard", "flashlight": "lamppost (~6m)", "bus_stop_sign": "sign on pole (~3.7m)",
}
REF_NOTE_RU = {
    "bus_stop": "неполный эталон: отсутствует дальняя от проезжей части сторона остановки; часть эталона — результат склейки 2 сканов, местами наложение неточное",
    "information_sign": "неполный эталон: заснята только верхняя ~2/3 высоты столба и одна сторона — сравнение корректно только в этой области (не ошибка масштаба/поворота)",
    "bench": "неполный эталон: отсутствует задняя часть спинки лавочки и ножки",
    "bollard": "неполный эталон: снята только одна сторона (~90° сектор отсутствует) — недостающая сторона учтена через детекцию дыр (DBSCAN)",
    "flashlight": "полный эталон",
    "bus_stop_sign": "полный эталон",
}
REF_NOTE_EN = {
    "bus_stop": "incomplete reference: the bus stop's far side (away from the road) wasn't scanned; part of the reference is a merge of 2 scans that don't align perfectly in places",
    "information_sign": "incomplete reference: only the upper ~2/3 of the pole height and one side were scanned — comparison is only valid in that region (not a scale/rotation error)",
    "bench": "incomplete reference: the bench's backrest (rear side) and legs are missing",
    "bollard": "incomplete reference: only one side was scanned (~90° sector missing) — the missing side is handled via gap detection (DBSCAN)",
    "flashlight": "full reference",
    "bus_stop_sign": "full reference",
}


def nn_spacing_median(points: np.ndarray) -> float:
    from scipy.spatial import cKDTree
    tree = cKDTree(points)
    d, _ = tree.query(points, k=2)
    return float(np.median(d[:, 1]))


def dbscan_mode_str(dbscan_cfg: dict, checkbox_checked: bool) -> str:
    if checkbox_checked:
        return "no DBSCAN"
    ft, eps, mp = dbscan_cfg["ft_default"], dbscan_cfg["eps_default"], dbscan_cfg["mp_default"]
    return f"ft{ft:g}/eps{eps:g}/mp{mp:g}"


def f_score(p: float, r: float) -> float:
    return 0.0 if (p + r) == 0 else 2 * p * r / (p + r)


# robust axis-aligned bbox: 0.5/99.5 percentile per axis, not true min/max, so one stray
# floating point (a real risk on noisier methods like vggt) doesn't blow up the reported size
BBOX_PCT_LO, BBOX_PCT_HI = 0.5, 99.5


def bbox_dims_cm(points_m: np.ndarray) -> tuple[float, float, float]:
    """(length, width, height) in cm from a point cloud in meters. Z is trusted as the world
    "up" axis (every reference/reconstruction here has already had the ground plane removed
    and been centered - see remove_ground_plane.py); the two horizontal extents (X, Y) aren't
    guaranteed to align with the object's own long/short axes (registration doesn't rotate to
    match), so they're just sorted: length = the longer footprint extent, width = the shorter."""
    lo = np.percentile(points_m, BBOX_PCT_LO, axis=0)
    hi = np.percentile(points_m, BBOX_PCT_HI, axis=0)
    extent_m = hi - lo
    height_cm = float(extent_m[2]) * 100.0
    horiz = sorted([float(extent_m[0]), float(extent_m[1])], reverse=True)
    length_cm, width_cm = horiz[0] * 100.0, horiz[1] * 100.0
    return length_cm, width_cm, height_cm


def load_reference(path) -> o3d.geometry.PointCloud:
    """Read a reference (LiDAR) cloud and drop exact-duplicate points before any
    metric touches it.

    The delivered clouds carry a lot of them - 10-26% of the points on these objects
    (bus_stop_001 41337 -> 31692, flashlight_003 22637 -> 16690, bollard_003 2818 ->
    2526) - left over where overlapping scan passes cover the same surface twice, at
    byte-identical coordinates. They are not extra information, but they do skew the
    metrics: completeness is averaged over target points, so a duplicated point is
    counted as many times as it appears, quietly weighting the score toward whatever
    the scanner happened to pass twice. Distances themselves are unaffected (a
    duplicate is its own nearest neighbour), so this only removes the double-counting.
    """
    pcd = o3d.io.read_point_cloud(str(path))
    pts = np.asarray(pcd.points)
    _, first_idx = np.unique(pts, axis=0, return_index=True)
    if len(first_idx) < len(pts):
        # select_by_index (rather than rebuilding from the array) so colors/normals survive
        pcd = pcd.select_by_index(np.sort(first_idx).tolist())
        print(f"       deduplicated reference: {len(pts)} -> {len(first_idx)} points "
              f"({100 * (1 - len(first_idx) / len(pts)):.1f}% exact duplicates removed)", flush=True)
    return pcd


RMSE_PATH = PROJECT_ROOT / "docs" / "tables" / "registration_rmse_from_aligned_clouds.json"
_rmse_table = json.loads(RMSE_PATH.read_text())["rows"] if RMSE_PATH.exists() else {}


def load_icp_rmse_mm(exp_id: str) -> tuple[float | None, float | None]:
    """Registration quality -> (rmse_mm, inlier_share_pct).

    Reported alongside F1 so a reader can tell "F1 differs because the method differs" from
    "F1 differs because this particular registration is worse" - it answers "maybe it's just
    badly aligned?" directly instead of leaving it to be inferred.

    Deliberately NOT each registration's own report.json: those describe superseded
    alignments for 14 of the 15 rows that have one (the aligned .ply was rewritten minutes
    to hours after transform.txt, so the stored inlier_rmse belongs to a cloud that is no
    longer on disk), and the other 9 report.json files are 0 bytes. The sidecar read here is
    measured from the same aligned .ply every accuracy number on this page uses, against the
    same reference, on the same 1 cm voxel grid and 3 cm threshold as F1@3cm - so the two
    columns describe the same clouds. See its "why" field.

    Nothing here reads or writes outputs/registrations/, which is deliberately read-only.
    """
    entry = _rmse_table.get(exp_id)
    if not entry:
        return None, None
    return entry.get("rmse_inlier_mm"), entry.get("inlier_share_pct")


def compute_rows() -> list[dict]:
    rows = []
    for page_id, cfg in MERGED_OBJECTS.items():
        ref_path = PROJECT_ROOT / cfg["ref"]
        print(f"[ref] {page_id}: loading {ref_path.name}", flush=True)
        ref = load_reference(ref_path)
        rpts = np.asarray(ref.points)
        median_spacing = nn_spacing_median(rpts)
        ref_spacing_cm = median_spacing * 100.0

        dbscan_cfg = {**DEFAULT_DBSCAN_SLIDERS, **cfg.get("dbscan", {})}
        checkbox_checked = cfg["checkbox_checked"]
        dbscan_mode = dbscan_mode_str(dbscan_cfg, checkbox_checked)

        capture = cfg["captures"][0]  # exactly one capture per object now
        for method_id in METHOD_ORDER:
            if method_id not in capture["methods"]:
                continue
            exp_id, rel_path = capture["methods"][method_id]
            src_path = PROJECT_ROOT / rel_path
            print(f"  [{page_id}/{method_id}] {exp_id} <- {src_path.name}", flush=True)
            src = o3d.io.read_point_cloud(str(src_path))
            raw_points = len(src.points)
            matched = src.voxel_down_sample(VOXEL_M)
            mpts = np.asarray(matched.points)
            icp_rmse_mm, icp_inlier_pct = load_icp_rmse_mm(exp_id)

            d_s2t_cm = np.asarray(matched.compute_point_cloud_distance(ref)) * 100.0
            d_t2s_cm = np.asarray(ref.compute_point_cloud_distance(matched)) * 100.0

            if checkbox_checked:
                kept_mask = np.ones(len(mpts), dtype=bool)  # honest mode: nothing excluded
            else:
                below_mask = d_s2t_cm <= FLOOR_CM
                gap_mask = np.zeros(len(mpts), dtype=bool)
                far_idx = np.where(d_s2t_cm > dbscan_cfg["ft_default"])[0]
                if len(far_idx) > 0:
                    far_pcd = o3d.geometry.PointCloud()
                    far_pcd.points = o3d.utility.Vector3dVector(mpts[far_idx])
                    labels = np.array(far_pcd.cluster_dbscan(eps=dbscan_cfg["eps_default"] / 100.0, min_points=dbscan_cfg["mp_default"]))
                    gap_mask[far_idx[labels >= 0]] = True
                kept_mask = ~gap_mask

            n_excluded = int((~kept_mask).sum())
            d_kept_cm = d_s2t_cm[kept_mask]
            acc_median = float(np.median(d_kept_cm)) if d_kept_cm.size else float("nan")
            comp_median = float(np.median(d_t2s_cm)) if d_t2s_cm.size else float("nan")
            length_cm, width_cm, height_cm = bbox_dims_cm(mpts)

            row = {
                "object_id": capture["id"],
                "shape_ru": SHAPE_RU[page_id], "shape_en": SHAPE_EN[page_id],
                "ref_spacing_cm": round(ref_spacing_cm, 3),
                "method": method_id,
                "accuracy_median_cm": round(acc_median, 2),
                "completeness_median_cm": round(comp_median, 2),
                "icp_rmse_mm": icp_rmse_mm,
                "icp_inlier_pct": icp_inlier_pct,
                "raw_points": raw_points,
                "matched_points": len(mpts),
                "raw_to_matched_ratio": round(raw_points / len(mpts), 2) if len(mpts) else None,
                "dbscan_mode": dbscan_mode,
                "page_id": page_id,
                "panel_key": f'{capture["id"]}__{method_id}',
                "n_excluded": n_excluded,
                "ref_note_ru": REF_NOTE_RU[page_id], "ref_note_en": REF_NOTE_EN[page_id],
                "length_cm": round(length_cm, 1), "width_cm": round(width_cm, 1), "height_cm": round(height_cm, 1),
            }
            metrics_log = []
            for t in THRESHOLDS_CM:
                acc_t = float(np.mean(d_kept_cm <= t)) if d_kept_cm.size else 0.0
                comp_t = float(np.mean(d_t2s_cm <= t)) if d_t2s_cm.size else 0.0
                f1_t = f_score(acc_t, comp_t)
                key = f"{t:g}cm"
                row[f"accuracy_{key}_pct"] = round(acc_t * 100, 1)
                row[f"completeness_{key}_pct"] = round(comp_t * 100, 1)
                row[f"f1_{key}_pct"] = round(f1_t * 100, 1)
                metrics_log.append(f"acc@{key}={acc_t*100:.1f}% comp@{key}={comp_t*100:.1f}% F1@{key}={f1_t*100:.1f}%")
            # delta F1@10cm - F1@3cm: separates a pure offset/bias (closes fast as the threshold
            # widens, small delta) from a real gap/missing-coverage problem (stays low even at
            # 10cm, large delta) - the distinction the three thresholds were added for in the
            # first place, made explicit instead of left for the reader to compute themselves.
            row["f1_delta_10_3_pct"] = round(row["f1_10cm_pct"] - row["f1_3cm_pct"], 1)

            rmse_str = f"{icp_rmse_mm:.1f}mm/{icp_inlier_pct:.0f}%" if icp_rmse_mm is not None else "N/A"
            print(f"          acc_med={acc_median:.2f}cm comp_med={comp_median:.2f}cm  " + " ".join(metrics_log) +
                  f"  ΔF1(10-3)={row['f1_delta_10_3_pct']:+.1f}pp  ICP_RMSE={rmse_str}  "
                  f"raw={raw_points} matched={len(mpts)} (x{row['raw_to_matched_ratio']})  "
                  f"L={length_cm:.0f}cm W={width_cm:.0f}cm H={height_cm:.0f}cm", flush=True)

            rows.append(row)
    return rows


def fmt_rmse(r: dict):
    """RMSE cell - measured from the same aligned cloud and reference as this row's F1."""
    v = r["icp_rmse_mm"]
    return "N/A" if v is None else v


def write_xlsx(rows: list[dict], path: Path, lang: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    thresh_keys = [f"{t:g}cm" for t in THRESHOLDS_CM]
    if lang == "ru":
        headers = ["object_id", "форма", "шаг реф., см", "длина, см", "ширина, см", "высота, см",
                   "метод", "accuracy median, см", "completeness median, см",
                   "RMSE выравнивания, мм", "точек в 3см, %"]
        f1_cols = []
        for k in thresh_keys:
            headers += [f"accuracy@{k}, %", f"completeness@{k}, %", f"F1@{k}, %"]
            f1_cols.append(len(headers))
        headers += ["ΔF1@10-3см, п.п.", "точек до вокселя", "точек после 1см", "raw/matched, ×",
                    "режим DBSCAN", "примечание по эталону"]
        shape_key, note_key = "shape_ru", "ref_note_ru"
    else:
        headers = ["object_id", "shape", "ref. spacing (cm)", "length (cm)", "width (cm)", "height (cm)",
                   "method", "accuracy median (cm)", "completeness median (cm)",
                   "alignment RMSE (mm)", "points within 3cm (%)"]
        f1_cols = []
        for k in thresh_keys:
            headers += [f"accuracy@{k} (%)", f"completeness@{k} (%)", f"F1@{k} (%)"]
            f1_cols.append(len(headers))
        headers += ["ΔF1@10-3cm (pp)", "raw points", "matched points (1cm voxel)", "raw/matched ratio",
                    "DBSCAN mode", "reference note"]
        shape_key, note_key = "shape_en", "ref_note_en"
    n_cols = len(headers)
    dbscan_col, note_col = n_cols - 1, n_cols

    wb = Workbook()
    ws = wb.active
    ws.title = "summary"
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")
    for c in ws[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    best_font = Font(bold=True)

    # winner (best F1@3cm) per object, computed up front - its bbox (length/width/height) is
    # what gets shown for the whole group, per "измерь по границам лучшей реконструкции".
    # 3cm (not 5/10) stays the tie-breaker since it's also what the DBSCAN gap mask floor uses.
    obj_groups: dict[str, list[int]] = {}
    for i, r in enumerate(rows):
        obj_groups.setdefault(r["object_id"], []).append(i)
    best_idx_by_obj = {obj_id: max(idxs, key=lambda i: rows[i]["f1_3cm_pct"]) for obj_id, idxs in obj_groups.items()}
    # separately, best-per-threshold within each object, for bolding each F1 column independently
    best_idx_by_obj_thresh = {
        (obj_id, k): max(idxs, key=lambda i: rows[i][f"f1_{k}_pct"])
        for obj_id, idxs in obj_groups.items() for k in thresh_keys
    }

    for r in rows:
        best = rows[best_idx_by_obj[r["object_id"]]]
        vals = [
            r["object_id"], r[shape_key], r["ref_spacing_cm"],
            best["length_cm"], best["width_cm"], best["height_cm"],
            r["method"],
            r["accuracy_median_cm"], r["completeness_median_cm"], fmt_rmse(r), r["icp_inlier_pct"],
        ]
        for k in thresh_keys:
            vals += [r[f"accuracy_{k}_pct"], r[f"completeness_{k}_pct"], r[f"f1_{k}_pct"]]
        vals += [r["f1_delta_10_3_pct"], r["raw_points"], r["matched_points"], r["raw_to_matched_ratio"],
                 r["dbscan_mode"], r[note_key]]
        ws.append(vals)

    # bold the best F1 per object group, independently per threshold column
    for (obj_id, k), best_i in best_idx_by_obj_thresh.items():
        col = f1_cols[thresh_keys.index(k)]
        ws.cell(row=best_i + 2, column=col).font = best_font

    # merge the per-object columns (object_id, shape, ref spacing, dimensions, DBSCAN mode,
    # reference note) down each object's row block - those repeat identically across its 4
    # method rows, only "метод"/"method" and the numeric accuracy columns vary per row.
    top_border = Border(top=Side(style="medium", color="9AA07A"))
    merge_cols = [1, 2, 3, 4, 5, 6, dbscan_col, note_col]
    for obj_id, idxs in obj_groups.items():
        first_row, last_row = idxs[0] + 2, idxs[-1] + 2  # +2: header row + 1-indexing
        for col in merge_cols:
            ws.merge_cells(start_row=first_row, start_column=col, end_row=last_row, end_column=col)
            top_cell = ws.cell(row=first_row, column=col)
            top_cell.alignment = Alignment(vertical="center", wrap_text=(col == note_col), horizontal=("left" if col in (1, 2, note_col) else "center"))
        for col in range(1, n_cols + 1):
            ws.cell(row=first_row, column=col).border = top_border

    widths = [20, 14, 12, 11, 11, 11, 12, 15, 17, 15, 13] + [13, 16, 10] * len(thresh_keys) + [14, 12, 15, 13, 14, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 32

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"Wrote {path.relative_to(PROJECT_ROOT)}")


def write_json(rows: list[dict], path: Path) -> None:
    """Per-page, per-panel exact metrics for the site to display verbatim."""
    pages: dict[str, dict] = {}
    for r in rows:
        page = pages.setdefault(r["page_id"], {"dbscan_mode": r["dbscan_mode"], "panels": {}})
        page["panels"][r["panel_key"]] = {
            **{f"{m}_{t:g}cm": r[f"{name}_{t:g}cm_pct"]
               for t in THRESHOLDS_CM
               for m, name in (("acc", "accuracy"), ("comp", "completeness"), ("f1", "f1"))},
            "acc_median_cm": r["accuracy_median_cm"],
            "comp_median_cm": r["completeness_median_cm"],
            "n_excluded": r["n_excluded"],
        }
    path.write_text(json.dumps({"source": "build_accuracy_f1_summary_table.py", "pages": pages}, indent=2))
    print(f"Wrote {path.relative_to(PROJECT_ROOT)}")


def main() -> None:
    rows = compute_rows()
    write_xlsx(rows, OUT_XLSX, "ru")
    write_xlsx(rows, OUT_XLSX_EN, "en")
    write_json(rows, OUT_JSON)


if __name__ == "__main__":
    main()
